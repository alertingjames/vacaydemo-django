# import datetime
import difflib
import os
import string
import urllib
from itertools import islice

import io
import requests
import xlrd
import re

from django.core import mail
from django.core.mail import send_mail, BadHeaderError, EmailMessage
from django.contrib import messages
# from _mysql_exceptions import DataError, IntegrityError
from django.template import RequestContext

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from django.core.mail import EmailMultiAlternatives

from django.core.files.storage import FileSystemStorage
import json
from django.contrib import auth
from django.http import HttpResponse, JsonResponse, HttpResponseRedirect, HttpResponseNotAllowed
from django.utils.datastructures import MultiValueDictKeyError
from django.views.decorators.cache import cache_control
from numpy import long
from openpyxl.styles import PatternFill

from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.fields import empty
from rest_framework.permissions import AllowAny
from time import gmtime, strftime
import time
from xlrd import XLRDError

from django.shortcuts import render, get_object_or_404, redirect
from django.views.decorators.csrf import csrf_protect, csrf_exempt
from django.contrib.sessions.backends.db import SessionStore
from django.contrib.sessions.models import Session
from django.contrib.auth.models import User, AnonymousUser
from django.conf import settings
from django import forms
import sys
from django.core.cache import cache
import random
from pbsonesignal import PybossaOneSignal

from pyfcm import FCMNotification

from motherwise.models import Member, Contact, Group, GroupMember, GroupConnect, Post, Comment, PostPicture, PostLike, Notification, Received, Sent, Replied, Conference
# from motherwise.serializers import

import pyrebase

config = {
    "apiKey": "AIzaSyATFKJgIovPNZY6MCUyMAM-UK0T5qWdqWU",
    "authDomain": "vacay-demo-1c8e4.firebaseapp.com",
    "databaseURL": "https://vacay-demo-1c8e4.firebaseio.com",
    "storageBucket": "vacay-demo-1c8e4.appspot.com"
}

firebase = pyrebase.initialize_app(config)


class UploadFileForm(forms.Form):
    file = forms.FileField()


from Crypto.Cipher import AES
from base64 import b64encode, b64decode

class Crypt:

    def __init__(self, salt='SlTKeYOpHygTYkP3'):
        self.salt = salt.encode('utf8')
        self.enc_dec_method = 'utf-8'

    def encrypt(self, str_to_enc, str_key):
        try:
            aes_obj = AES.new(str_key, AES.MODE_CFB, self.salt)
            hx_enc = aes_obj.encrypt(str_to_enc.encode('utf8'))
            mret = b64encode(hx_enc).decode(self.enc_dec_method)
            return mret
        except ValueError as value_error:
            if value_error.args[0] == 'IV must be 16 bytes long':
                raise ValueError('Encryption Error: SALT must be 16 characters long')
            elif value_error.args[0] == 'AES key must be either 16, 24, or 32 bytes long':
                raise ValueError('Encryption Error: Encryption key must be either 16, 24, or 32 characters long')
            else:
                raise ValueError(value_error)

    def decrypt(self, enc_str, str_key):
        try:
            aes_obj = AES.new(str_key.encode('utf8'), AES.MODE_CFB, self.salt)
            str_tmp = b64decode(enc_str.encode(self.enc_dec_method))
            str_dec = aes_obj.decrypt(str_tmp)
            mret = str_dec.decode(self.enc_dec_method)
            return mret
        except ValueError as value_error:
            if value_error.args[0] == 'IV must be 16 bytes long':
                raise ValueError('Decryption Error: SALT must be 16 characters long')
            elif value_error.args[0] == 'AES key must be either 16, 24, or 32 bytes long':
                raise ValueError('Decryption Error: Encryption key must be either 16, 24, or 32 characters long')
            else:
                raise ValueError(value_error)


def encrypt(info):
    crpt = Crypt()
    test_key = 'MyKey4TestingYnP'
    result = crpt.encrypt(info, test_key)
    return result


def decrypt(info):
    crpt = Crypt()
    test_key = 'MyKey4TestingYnP'
    result = crpt.decrypt(info, test_key)
    return result



def index(request):
    # try:
    #     if request.session['adminID'] != 0:
    #         return redirect('/home')
    # except KeyError:
    #     print('no session')

    try:
        if request.session['memberID'] != '' and request.session['memberID'] != 0:
            member_id = request.session['memberID']
            members = Member.objects.filter(id=member_id)
            if members.count() == 0:
                return render(request, 'mothers/login.html')
            member = members[0]
            if member.cohort == 'admin':
                return render(request, 'mothers/login.html')
            if member.photo_url == '' or member.cohort == '' or member.phone_number == '':
                return render(request, 'mothers/register_profile.html', {'member':member})
            elif member.address == '' or member.city == '':
                return  render(request, 'mothers/location_picker.html', {'address':member.address})
            else:
                return redirect('/users/home')
    except KeyError:
        print('no session')
    return render(request, 'mothers/login.html')

    # return redirect('/users/')




def admin(request):
    # return redirect('/logout')
    try:
        if request.session['adminID'] != '' and request.session['adminID'] != 0:
            return redirect('/home')
    except KeyError:
        print('no session')
    return render(request, 'motherwise/admin.html')

def adminsignuppage(request):
    return render(request, 'motherwise/adminsignup.html')

def adminloginpage(request):
    return render(request, 'motherwise/admin.html')

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def adminSignup(request):
    if request.method == 'POST':
        email = request.POST.get('email', '')
        name = request.POST.get('name', '')
        password = request.POST.get('password', '')
        phone_number = request.POST.get('phone_number', '')
        playerID = request.POST.get('playerID', '')

        members = Member.objects.filter(email=email, cohort='admin')
        count = members.count()
        if count ==0:
            member = Member()
            member.admin_id = '0'
            member.email = email
            # member.name = name
            member.name = 'VaCay Admin'
            member.password = password
            member.phone_number = phone_number
            member.photo_url = settings.URL + '/static/images/manager.jpg'
            member.cohort = 'admin'
            member.registered_time = str(int(round(time.time() * 1000)))
            member.playerID = playerID
            member.save()

            request.session['adminID'] = member.pk

            return redirect('/home')

        else:
            return redirect('/logout')


def adminhome(request):
    import datetime
    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    members = Member.objects.filter(admin_id=adminID).order_by('-id')
    for member in members:
        if member.registered_time != '':
            member.registered_time = datetime.datetime.fromtimestamp(float(int(member.registered_time)/1000)).strftime("%b %d, %Y")
    users, range = get_all_member_data(members)

    groups = Group.objects.filter(member_id=admin.pk).order_by('-id')

    return render(request, 'motherwise/adminhome.html', {'me':admin,'users':users, 'range': range, 'current': 1, 'groups':groups})


def get_all_member_data(members):
    i = 0
    memberList = []
    for member in members:
        i = i + 1
        if i <= 25:
            memberList.append(member)
    r = int(members.count() / 25)
    m = members.count() % 25
    if m > 0:
        r = r + 2
    else:
        r = r + 1

    return memberList, range(r)



def adminlogout(request):
    request.session['adminID'] = 0
    request.session['selected_option'] = ''
    request.session['selected_member_list'] = []
    return render(request, 'motherwise/admin.html')


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def adminLogin(request):
    if request.method == 'POST':
        email = request.POST.get('email', '')
        password = request.POST.get('password', '')
        playerID = request.POST.get('playerID', '')

        members = Member.objects.filter(email=email, password=password, cohort='admin')
        if members.count() > 0:
            member = members[0]
            if playerID != '':
                member.playerID = playerID
                member.save()
            request.session['adminID'] = member.pk
            return redirect('/home')
        else:
            return render(request, 'motherwise/result.html',
                          {'response': 'You don\'t have any permission to access this site. Try again with another credential.'})

def export_xlsx_member(request):
    import openpyxl
    from openpyxl.utils import get_column_letter
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=member_template.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Members"

    row_num = 0

    columns = [
        (u"Name", 30),
        (u"E-mail", 50),
        (u"Phone Number (+x xxx xxx xxxx)", 30),
        (u"Group Name(Refer to Sheet2)", 30),
        (u"City Name", 30),
        (u"Address", 80),
    ]

    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]

        # set column width
        ws.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]

    ws2=wb.create_sheet(title='Sheet2')

    ws2.column_dimensions["A"].width = 20
    my_color = openpyxl.styles.colors.Color(rgb='00ffaa02')
    my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_color)

    ws2['A1'].fill = my_fill
    ws2['A1'] = 'Group Names'
    ws2['A2'] = 'E81'
    ws2['A3'] = 'E83'
    ws2['A4'] = 'E84'
    ws2['A5'] = 'E86'
    ws2['A6'] = 'E87'
    ws2['A7'] = 'S82'
    ws2['A8'] = 'S85'
    ws2['A9'] = 'S88'
    ws2['A10'] = 'E(v)89'
    ws2['A11'] = 'E(v)90'
    ws2['A12'] = 'S(v)91'
    ws2['A13'] = 'E(v)92'
    ws2['A14'] = 'E(v)93'
    ws2['A15'] = 'S(v)94'
    ws2['A16'] = 'VACAY Leaders'

    wb.save(response)
    return response


def import_view_member(request):
    if request.method == "POST":
        form = UploadFileForm(request.POST,
                              request.FILES)
    else:
        form = UploadFileForm()

    return render(
        request,
        'motherwise/upload_form_member.html',
        {
            'form': form,
            'title': 'Load Data',
            'header': 'Upload Members From File'
        })


def import_member_data(request):
    if request.method == "POST":
        form = UploadFileForm(request.POST,
                              request.FILES)

        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']

        if form.is_valid():
            input_excel = request.FILES['file']
            try:
                book = xlrd.open_workbook(file_contents=input_excel.read())
                sheet = book.sheet_by_index(0)

                for r in range(1, sheet.nrows):
                    name = sheet.cell(r, 0).value
                    email = sheet.cell(r, 1).value
                    phone_number = sheet.cell(r, 2).value
                    group_name = sheet.cell(r, 3).value
                    city = sheet.cell(r, 4).value
                    address = sheet.cell(r, 5).value

                    name = name.strip()
                    email = email.strip()
                    city = city.strip()
                    address = address.strip()

                    members = Member.objects.filter(email=email)
                    if members.count() > 0:
                        continue
                    member = Member()
                    member.admin_id = adminID
                    member.name = name
                    member.email = email
                    member.password = generateRandomPassword()
                    member.phone_number = str(phone_number).replace('.0','').strip()
                    # group_number = str(group_number).replace('.0','')
                    # try:
                    #   val = int(group_number)
                    #   member.cohort = 'Group-' + str(val)
                    # except ValueError:
                    #   print("That's not an int!")
                    member.cohort = str(group_name).replace('.0','').strip()
                    member.city = city
                    member.address = address
                    member.lat = '0'
                    member.lng = '0'
                    member.save()

                    admin = Member.objects.get(id=adminID)

                    groupText = ''
                    if member.cohort != '':
                        groupText = '<br>Group: ' + member.cohort

                    title = 'Invitation for VaCay Community'
                    subject = 'VaCay Community'
                    message = 'Dear ' + member.name + ',<br><br>Welcome to VaCay\'s virtual community!<br><br>VaCay is an opportunity to connect and reconnect with other VaCay families.<br>'
                    message = message + 'You can post articles, share pregnancy and new baby tips, watch videos, and chat directly with other moms. You\'ll also stay up-to-date on all the new programs and special events VaCay has to offer!<br><br>'
                    message = message + settings.URL + '/users' + '<br><br>We are providing you with your initial login information as follows:<br><br>'
                    message = message + 'E-mail: ' + member.email + ' (your email)<br>Password: ' + member.password + groupText + '<br><br>'

                    message = message + '***By signing up to VaCay, you are agreeing to not engage in any type of: ***<br>'
                    message = message + '        · hate speech<br>'
                    message = message + '        · cyberbullying<br>'
                    message = message + '        · solicitation and/or selling of goods or services<br>'
                    message = message + '        · posting content inappropriate for our diverse community including but not limited to political<br>'
                    message = message + '    or religious views<br><br>'
                    message = message + 'We want VaCay to be a safe place for support and inspiration. Help us foster this community and please respect everyone on VaCay.<br><br>'
                    message = message + 'Please watch this video to see how to login: https://vimeo.com/430742850<br><br>'
                    message = message + 'If you have any question, please contact us:<br><br>'

                    message = message + '   E-mail: ' + settings.ADMIN_EMAIL + '<br>   Phone number: ' + '720-504-4624<br><br>'
                    message = message + '<a href=\'' + settings.URL + '/users' + '\' target=\'_blank\'>Join website</a><br><br>'
                    message = message + 'Sincerely<br><br>VaCay Team'

                    message = message + '<br><br>'

                    groupText2 = ''
                    if member.cohort != '':
                        groupText2 = '<br>Grupo: ' + member.cohort

                    title2 = 'Invitación para la comunidad de VaCay: El Nido'
                    message2 = 'Querida ' + member.name + ',<br><br>¡Bienvenida al \"Nido\": la comunidad virtual de VaCay!<br><br>El Nido es una oportunidad para conectarse y reconectarse con otras madres de VaCay.<br>'
                    message2 = message2 + 'Puede publicar artículos, compartir consejos sobre embarazo y nuevos bebés, ver videos y chatear directamente con otras madres. ¡También se mantendrá al tanto sobre todos los nuevos programas y eventos especiales que VaCay tiene para ofrecer!<br><br>'
                    message2 = message2 + settings.URL + '/users' + '<br><br>Le proporcionamos su información de inicio de la siguiente manera:<br><br>'
                    message2 = message2 + 'Correo electrónico: ' + member.email + ' (Tu correo electrónico)<br>Contraseña: ' + member.password + groupText2 + '<br><br>'

                    message2 = message2 + '***Al suscribirse al Nido, acepta no participar en ningún tipo de: ***<br>'
                    message2 = message2 + '        · El discurso del odio<br>'
                    message2 = message2 + '        · Ciberacoso<br>'
                    message2 = message2 + '        · Solicitud y/o venta de bienes o servicios<br>'
                    message2 = message2 + '        · Publicar contenido inapropiado para nuestra diversa comunidad, incluidos, entre otros, puntos de vista políticos o religiosos<br><br>'
                    message2 = message2 + 'Queremos que El Nido sea un lugar seguro para apoyo e inspiración. Por favor ayúdanos a fomentar esta comunidad y por favor respeta a todos en El Nido.<br><br>'
                    message2 = message2 + 'Mire este video para ver cómo iniciar sesión: https://vimeo.com/430742850<br><br>'
                    message2 = message2 + 'Si usted tiene cualquier pregunta, por favor póngase en contacto con nosotros:<br><br>'

                    message2 = message2 + '   Correo electrónico: ' + settings.ADMIN_EMAIL + '<br>   Número de teléfono: ' + '720-504-4624<br><br>'
                    message2 = message2 + '<a href=\'' + settings.URL + '/users' + '\' target=\'_blank\'>Unirse al sitio web</a><br><br>'
                    message2 = message2 + 'Sinceramente,<br><br>el Equipo de VaCay'

                    from_email = admin.email
                    to_emails = []
                    to_emails.append(member.email)
                    send_mail_message0(from_email, to_emails, title, subject, message, title2, message2)

                return redirect('/home')

            except XLRDError:
                return render(request, 'motherwise/upload_form_member.html', {'note': 'invalid_file'})

            except IOError:
                return render(request, 'motherwise/upload_form_member.html', {'note': 'invalid_file'})
            except IndexError:
                return render(request, 'motherwise/upload_form_member.html', {'note': 'invalid_file'})
            # except DataError:
            #     return HttpResponse('Invalid file!')
        else:
            return render(request, 'motherwise/upload_form_member.html', {'note': 'invalid_file'})


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def add_member(request):
    if request.method == 'POST':
        name = request.POST.get('name', '')
        email = request.POST.get('email', '')
        phone_number = request.POST.get('phone_number', '')
        cohort = request.POST.get('cohort', '')

        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']

        members = Member.objects.filter(email=email)
        if members.count() > 0:
            return render(request, 'motherwise/result.html',
                          {'response': 'This member has already been registered.'})
        member = Member()
        member.admin_id = adminID
        member.name = name
        member.email = email
        member.password = generateRandomPassword()
        member.phone_number = str(phone_number).replace('.0','')
        member.cohort = cohort
        member.lat = '0'
        member.lng = '0'
        member.save()

        admin = Member.objects.get(id=adminID)

        groupText = ''
        if member.cohort != '':
            groupText = '<br>Group: ' + member.cohort

        title = 'Invitation for VaCay Community'
        subject = 'VaCay Community'
        message = 'Dear ' + member.name + ',<br><br>Welcome to VaCay\'s virtual community!<br><br>VaCay is an opportunity to connect and reconnect with other VaCay families.<br>'
        message = message + 'You can post articles, share pregnancy and new baby tips, watch videos, and chat directly with other moms. You\'ll also stay up-to-date on all the new programs and special events VaCay has to offer!<br><br>'
        message = message + settings.URL + '/users' + '<br><br>We are providing you with your initial login information as follows:<br><br>'
        message = message + 'E-mail: ' + member.email + ' (your email)<br>Password: ' + member.password + groupText + '<br><br>'

        message = message + '***By signing up to VaCay, you are agreeing to not engage in any type of: ***<br>'
        message = message + '        · hate speech<br>'
        message = message + '        · cyberbullying<br>'
        message = message + '        · solicitation and/or selling of goods or services<br>'
        message = message + '        · posting content inappropriate for our diverse community including but not limited to political<br>'
        message = message + '    or religious views<br><br>'
        message = message + 'We want VaCay to be a safe place for support and inspiration. Help us foster this community and please respect everyone on VaCay.<br><br>'
        message = message + 'Please watch this video to see how to login: https://vimeo.com/430742850<br><br>'
        message = message + 'If you have any question, please contact us:<br><br>'

        message = message + '   E-mail: ' + settings.ADMIN_EMAIL + '<br>   Phone number: ' + '720-504-4624<br><br>'
        message = message + '<a href=\'' + settings.URL + '/users' + '\' target=\'_blank\'>Join website</a><br><br>'
        message = message + 'Sincerely<br><br>VaCay Team'

        message = message + '<br><br>'

        groupText2 = ''
        if member.cohort != '':
            groupText2 = '<br>Grupo: ' + member.cohort

        title2 = 'Invitación para la comunidad de VaCay: El Nido'
        message2 = 'Querida ' + member.name + ',<br><br>¡Bienvenida al \"Nido\": la comunidad virtual de VaCay!<br><br>El Nido es una oportunidad para conectarse y reconectarse con otras madres de VaCay.<br>'
        message2 = message2 + 'Puede publicar artículos, compartir consejos sobre embarazo y nuevos bebés, ver videos y chatear directamente con otras madres. ¡También se mantendrá al tanto sobre todos los nuevos programas y eventos especiales que VaCay tiene para ofrecer!<br><br>'
        message2 = message2 + settings.URL + '/users' + '<br><br>Le proporcionamos su información de inicio de la siguiente manera:<br><br>'
        message2 = message2 + 'Correo electrónico: ' + member.email + ' (Tu correo electrónico)<br>Contraseña: ' + member.password + groupText2 + '<br><br>'

        message2 = message2 + '***Al suscribirse al Nido, acepta no participar en ningún tipo de: ***<br>'
        message2 = message2 + '        · El discurso del odio<br>'
        message2 = message2 + '        · Ciberacoso<br>'
        message2 = message2 + '        · Solicitud y/o venta de bienes o servicios<br>'
        message2 = message2 + '        · Publicar contenido inapropiado para nuestra diversa comunidad, incluidos, entre otros, puntos de vista políticos o religiosos<br><br>'
        message2 = message2 + 'Queremos que El Nido sea un lugar seguro para apoyo e inspiración. Por favor ayúdanos a fomentar esta comunidad y por favor respeta a todos en El Nido.<br><br>'
        message2 = message2 + 'Mire este video para ver cómo iniciar sesión: https://vimeo.com/430742850<br><br>'
        message2 = message2 + 'Si usted tiene cualquier pregunta, por favor póngase en contacto con nosotros:<br><br>'

        message2 = message2 + '   Correo electrónico: ' + settings.ADMIN_EMAIL + '<br>   Número de teléfono: ' + '720-504-4624<br><br>'
        message2 = message2 + '<a href=\'' + settings.URL + '/users' + '\' target=\'_blank\'>Unirse al sitio web</a><br><br>'
        message2 = message2 + 'Sinceramente,<br><br>el Equipo de VaCay'

        from_email = admin.email
        to_emails = []
        to_emails.append(member.email)
        send_mail_message0(from_email, to_emails, title, subject, message, title2, message2)

        return redirect('/home')


def send_mail_message0(from_email, to_emails, title, subject, message, title2, message2):
    html =  """\
                <html>
                    <head></head>
                    <body>
                        <a href="#"><img src="https://www.vacaydemo.com/static/images/vacaylogo.jpg" style="width:120px;height:120px;border-radius: 8%; margin-left:25px;"/></a>
                        <h2 style="margin-left:10px; color:#02839a;">{title}</h2>
                        <div style="font-size:14px; white-space: pre-line; word-wrap: break-word;">
                            {mes}
                        </div>

                    </body>
                </html>
            """
    # html = html.format(title=title, mes=message, title2=title2, mes2=message2)
    html = html.format(title=title, mes=message)

    msg = EmailMultiAlternatives(subject, '', from_email, to_emails)
    msg.attach_alternative(html, "text/html")
    msg.send(fail_silently=False)



def send_mail_message(from_email, to_emails, title, subject, message):
    html =  """\
                <html>
                    <head></head>
                    <body>
                        <a href="#"><img src="https://www.vacaydemo.com/static/images/vacaylogo.jpg" style="width:120px;height:120px;border-radius: 8%; margin-left:25px;"/></a>
                        <h2 style="margin-left:10px; color:#02839a;">{title}</h2>
                        <div style="font-size:14px; white-space: pre-line; word-wrap: break-word;">
                            {mes}
                        </div>
                    </body>
                </html>
            """
    html = html.format(title=title, mes=message)

    msg = EmailMultiAlternatives(subject, '', from_email, to_emails)
    msg.attach_alternative(html, "text/html")
    msg.send(fail_silently=False)



def generateRandomPassword():
    import strgen
    randomString = strgen.StringGenerator("[\w\d]{10}").render()
    return randomString



@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def delete_member(request):
    if request.method == 'GET':
        member_id = request.GET['member_id']

        members = Member.objects.filter(id=member_id)

        fs = FileSystemStorage()

        if members.count() > 0:
            member = members[0]
            if member.photo_url != '' and '/static/images/ic_profile.png' not in member.photo_url:
                fname = member.photo_url.replace(settings.URL + '/media/', '')
                fs.delete(fname)
            member.delete()

            contacts = Contact.objects.filter(member_id=member_id)
            for contact in contacts:
                contact.delete()

            gms = GroupMember.objects.filter(member_id=member_id)
            for gm in gms:
                groups = Group.objects.filter(id=gm.group_id)
                if groups.count() > 0:
                    group = groups[0]
                    if int(group.member_count) > 0:
                        group.member_count = int(group.member_count) - 1
                gm.delete()

            gcs = GroupConnect.objects.filter(member_id=member_id)
            for gc in gcs:
                gc.delete()

            posts = Post.objects.filter(member_id=member_id)
            for post in posts:
                pps = PostPicture.objects.filter(post_id=post.pk)
                for pp in pps:
                    if pp.picture_url != '':
                       fname = pp.picture_url.replace(settings.URL + '/media/', '')
                       fs.delete(fname)
                    pp.delete()
                post.delete()

            comments = Comment.objects.filter(member_id=member_id)
            for comment in comments:
                comment.delete()

            pls = PostLike.objects.filter(member_id=member_id)
            for pl in pls:
                pl.delete()

            notis = Notification.objects.filter(member_id=member_id)
            for noti in notis:
                noti.delete()

            notis = Notification.objects.filter(sender_id=member_id)
            for noti in notis:
                noti.delete()

            return redirect('/home')
        else:
            return render(request, 'motherwise/result.html',
                          {'response': 'This member doesn\'t exist. Please refresh the site.'})



def active_members(request):
    import datetime
    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    members = Member.objects.filter(admin_id=adminID).order_by('-id')
    for member in members:
        if member.registered_time != '':
            member.registered_time = datetime.datetime.fromtimestamp(float(int(member.registered_time)/1000)).strftime("%b %d, %Y")
    i = 0
    memberList = []
    for member in members:
        i = i + 1
        if i <= 25:
            if member.registered_time != '':
                memberList.append(member)
    r = int(len(memberList) / 25)
    m = len(memberList) % 25
    if m > 0:
        r = r + 2
    else:
        r = r + 1
    return render(request, 'motherwise/adminhome.html', {'me':admin,'users':memberList, 'range': range(r), 'current': 1, 'title':'Active members'})


def inactive_members(request):
    import datetime
    try:
        if request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    members = Member.objects.filter(admin_id=adminID).order_by('-id')
    for member in members:
        if member.registered_time != '':
            member.registered_time = datetime.datetime.fromtimestamp(float(int(member.registered_time)/1000)).strftime("%b %d, %Y")
    i = 0
    memberList = []
    for member in members:
        i = i + 1
        if i <= 25:
            if member.registered_time == '':
                memberList.append(member)
    r = int(len(memberList) / 25)
    m = len(memberList) % 25
    if m > 0:
        r = r + 2
    else:
        r = r + 1
    return render(request, 'motherwise/adminhome.html', {'me':admin,'users':memberList, 'range': range(r), 'current': 1, 'title':'Inactive members'})


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def message_to_selected_members(request):

    import datetime

    if request.method == 'POST':

        ids = request.POST.getlist('users2[]')

        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        try:
            option = request.POST.get('option','')
            if option == 'private_chat':
                memberList = []
                memberIdList = []
                for member_id in ids:
                    members = Member.objects.filter(id=int(member_id))
                    if members.count() > 0:
                        member = members[0]
                        memberList.append(member)
                        memberIdList.append(member.pk)

                contacts = update_admin_contact(admin, "")

                if len(memberList) > 0:
                    request.session['selected_option'] = option
                    request.session['selected_member_list'] = memberIdList
                    return render(request, 'motherwise/chat.html', {'members':memberList, 'me': admin, 'friend':memberList[0], 'contacts':contacts})
                else:
                    return redirect('/home')

            elif option == 'group_chat':
                memberList = []
                memberList2 = []
                memberIdList = []
                for member_id in ids:
                    members = Member.objects.filter(id=int(member_id))
                    if members.count() > 0:
                        member = members[0]
                        memberList.append(member)
                        memberIdList.append(member.pk)
                if len(memberList) > 0:
                    request.session['selected_option'] = option
                    request.session['selected_member_list'] = memberIdList
                    groups = Group.objects.filter(member_id=admin.pk).order_by('-id')
                    for group in groups:
                        group.created_time = datetime.datetime.fromtimestamp(float(int(group.created_time)/1000)).strftime("%b %d, %Y %H:%M")
                        group.last_connected_time = datetime.datetime.fromtimestamp(float(int(group.last_connected_time)/1000)).strftime("%b %d, %Y %H:%M")
                    latestGroupMemberList = []
                    latest_group = None
                    if groups.count() > 0:
                        latest_group = groups[0]
                        gMembers = GroupMember.objects.filter(group_id=latest_group.pk)
                        for gMember in gMembers:
                            members = Member.objects.filter(id=gMember.member_id)
                            if members.count() > 0:
                                latestGroupMemberList.append(members[0])
                        for memb in memberList:
                            gms = GroupMember.objects.filter(group_id=latest_group.pk, member_id=memb.pk)
                            if gms.count() == 0: memberList2.append(memb)
                    else:
                        memberList2 = memberList
                    gcs = GroupConnect.objects.filter(member_id=admin.pk).order_by('-id')
                    recents = []
                    for gc in gcs:
                        gs = Group.objects.filter(id=gc.group_id)
                        if gs.count() > 0:
                            group = gs[0]
                            group.created_time = datetime.datetime.fromtimestamp(float(int(group.created_time)/1000)).strftime("%b %d, %Y %H:%M")
                            group.last_connected_time = datetime.datetime.fromtimestamp(float(int(group.last_connected_time)/1000)).strftime("%b %d, %Y %H:%M")
                            recents.append(group)

                    return render(request, 'motherwise/groups.html', {'members':memberList2, 'group':latest_group, 'groups': groups, 'group_members':latestGroupMemberList, 'recents':recents})
                else:
                    return redirect('/home')


        except KeyError:
            print('no such key')

        message = request.POST.get('message', '')

        for member_id in ids:
            members = Member.objects.filter(id=int(member_id))
            if members.count() > 0:
                member = members[0]

                notification = Notification()
                notification.member_id = member.pk
                notification.sender_id = admin.pk
                notification.message = message
                notification.notified_time = str(int(round(time.time() * 1000)))
                notification.save()

                rcv = Received()
                rcv.member_id = member.pk
                rcv.sender_id = admin.pk
                rcv.noti_id = notification.pk
                rcv.save()

                snt = Sent()
                snt.member_id = member.pk
                snt.sender_id = admin.pk
                snt.noti_id = notification.pk
                snt.save()

                title = 'VaCay Community'
                subject = 'You\'ve received a message from VaCay Community'
                msg = 'Dear ' + member.name + ', You\'ve received a message from VaCay Community. The message is as following:<br><br>'
                msg = msg + message
                msg = msg + '<br><br><a href=\'' + settings.URL + '/users/notifications?noti_id=' + str(notification.pk) + '\' target=\'_blank\'>Join website</a>'

                from_email = admin.email
                to_emails = []
                to_emails.append(member.email)
                send_mail_message(from_email, to_emails, title, subject, msg)

                msg = member.name + ', You\'ve received a message from VaCay Community.\nThe message is as following:\n' + message

                ##########################################################################################################################################################################

                db = firebase.database()
                data = {
                    "msg": message,
                    "date":str(int(round(time.time() * 1000))),
                    "sender_id": str(admin.pk),
                    "sender_name": admin.name,
                    "sender_email": admin.email,
                    "sender_photo": admin.photo_url,
                    "role": "admin",
                    "type": "message",
                    "id": str(notification.pk),
                    "mes_id": str(notification.pk)
                }

                db.child("notify").child(str(member.pk)).push(data)
                db.child("notify2").child(str(member.pk)).push(data)

                sendFCMPushNotification(member.pk, admin.pk, message)

                #################################################################################################################################################################################

                if member.playerID != '':
                    playerIDList = []
                    playerIDList.append(member.playerID)
                    url = '/users/notifications?noti_id=' + str(notification.pk)
                    send_push(playerIDList, msg, url)

        members = Member.objects.filter(admin_id=adminID).order_by('-id')
        users, range = get_all_member_data(members)

        return render(request, 'motherwise/adminhome.html', {'me':admin,'users':users, 'range': range, 'current': 1, 'notify':'message_sent'})

    else:
        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        memberIdList = []
        try:
            memberIdList = request.session['selected_member_list']
        except KeyError:
            print('No key')

        memberList = []
        for member_id in memberIdList:
            members = Member.objects.filter(id=member_id)
            if members.count() > 0:
                member = members[0]
                memberList.append(member)
        selectedOption = request.session['selected_option']

        contacts = update_admin_contact(admin, "")

        if len(memberList) == 0:
            return render(request, 'motherwise/result.html',
                          {'response': 'The members don\'t exist.'})

        if len(memberList) > 0:
            if selectedOption == 'private_chat':
                return render(request, 'motherwise/chat.html', {'members':memberList, 'me': admin, 'friend':memberList[0], 'contacts':contacts})
            elif selectedOption == 'group_chat':
                groups = Group.objects.filter(member_id=admin.pk).order_by('-id')
                for group in groups:
                    group.created_time = datetime.datetime.fromtimestamp(float(int(group.created_time)/1000)).strftime("%b %d, %Y %H:%M")
                    group.last_connected_time = datetime.datetime.fromtimestamp(float(int(group.last_connected_time)/1000)).strftime("%b %d, %Y %H:%M")
                latestGroupMemberList = []
                latest_group = None
                if groups.count() > 0:
                    latest_group = groups[0]
                    gMembers = GroupMember.objects.filter(group_id=latest_group.pk)
                    for gMember in gMembers:
                        members = Member.objects.filter(id=gMember.member_id)
                        if members.count() > 0:
                            latestGroupMemberList.append(members[0])
                gcs = GroupConnect.objects.filter(member_id=admin.pk).order_by('-id')
                recents = []
                for gc in gcs:
                    gs = Group.objects.filter(id=gc.group_id)
                    if gs.count() > 0:
                        group = gs[0]
                        group.created_time = datetime.datetime.fromtimestamp(float(int(group.created_time)/1000)).strftime("%b %d, %Y %H:%M")
                        group.last_connected_time = datetime.datetime.fromtimestamp(float(int(group.last_connected_time)/1000)).strftime("%b %d, %Y %H:%M")
                        recents.append(group)
                return render(request, 'motherwise/groups.html', {'members':memberList, 'group':latest_group, 'groups': groups, 'group_members':latestGroupMemberList, 'recents':recents})
            else:
                return redirect('/home')
        else:
            return redirect('/home')



def to_page(request):
    index = request.GET['index']
    page = request.GET['page']

    import datetime

    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    i = 0
    if page == 'all_members':
        if int(index) == 1:
            return redirect('/home')
        userList = []
        users = Member.objects.filter(admin_id=admin.pk).order_by('-id')
        for user in users:
            if user.registered_time != '':
                user.registered_time = datetime.datetime.fromtimestamp(float(int(user.registered_time)/1000)).strftime("%b %d, %Y")
        for user in users:
            i = i + 1
            if i > 25 * (int(index) - 1) and i <= 25 * int(index):
                userList.append(user)
        r = int(users.count() / 25)
        r = r + 2
        return render(request, 'motherwise/adminhome.html', {'users':userList, 'range': range(r), 'current': index})

    elif page == 'active_members':
        if int(index) == 1:
            return redirect('/active_members')
        userList = []
        users = Member.objects.filter(admin_id=admin.pk).order_by('-id')
        for user in users:
            if user.registered_time != '':
                user.registered_time = datetime.datetime.fromtimestamp(float(int(user.registered_time)/1000)).strftime("%b %d, %Y")
        for user in users:
            i = i + 1
            if i > 25 * int(index - 1) and i <= 25 * int(index):
                if user.registered_time != '':
                    userList.append(user)
        r = int(users.count() / 25)
        r = r + 2
        return render(request, 'motherwise/adminhome.html', {'users':userList, 'range': range(r), 'current': index})

    elif page == 'inactive_members':
        if int(index) == 1:
            return redirect('/inactive_members')
        userList = []
        users = Member.objects.filter(admin_id=admin.pk).order_by('-id')
        for user in users:
            if user.registered_time != '':
                user.registered_time = datetime.datetime.fromtimestamp(float(int(user.registered_time)/1000)).strftime("%b %d, %Y")
        for user in users:
            i = i + 1
            if i > 25 * int(index - 1) and i <= 25 * int(index):
                if user.registered_time == '':
                    userList.append(user)
        r = int(users.count() / 25)
        r = r + 2
        return render(request, 'motherwise/adminhome.html', {'users':userList, 'range': range(r), 'current': index})


def to_previous(request):
    index = request.GET['index']
    page = request.GET['page']

    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    if page == 'all_members':
        if int(index) == 1:
            return redirect('/home')
    elif page == 'active_members':
        if int(index) == 1:
            return redirect('/active_members')
    elif page == 'inactive_members':
        if int(index) == 1:
            return redirect('/inactive_members')

    index = int(index) - 1
    return redirect('/to_page?index=' + str(index) + '&page=' + page)


def to_next(request):
    index = request.GET['index']
    page = request.GET['page']

    try:
        if request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    count = 0
    if page == 'all_members':
        users = Member.objects.filter(admin_id=admin.pk).order_by('-id')
        count = users.count()

    elif page == 'active_members':
        users = Member.objects.filter(admin_id=admin.pk).order_by('-id')
        userList = []
        for user in users:
            if user.registered_time != '':
                userList.append(user)
        count = len(userList)

    elif page == 'inactive_members':
        users = Member.objects.filter(admin_id=admin.pk).order_by('-id')
        userList = []
        for user in users:
            if user.registered_time == '':
                userList.append(user)
        count = len(userList)

    r = int(count / 25)
    m = count % 25
    if m > 0:
        r = r + 2
    else:
        r = r + 1
    if int(index) < r - 1:
        index = int(index) + 1
    return redirect('/to_page?index=' + str(index) + '&page=' + page)


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def do_cohort(request):

    import datetime

    if request.method == 'POST':
        try:
            cohort = request.POST.get('cohort','')
            option = request.POST.get('option','')
        except AssertionError:
            return redirect('/home')

        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        if cohort == '':
            return render(request, 'motherwise/result.html',
                          {'response': 'Please choose a cohort.'})

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        members = Member.objects.filter(admin_id=adminID).order_by('-id')
        for member in members:
            if member.registered_time != '':
                member.registered_time = datetime.datetime.fromtimestamp(float(int(member.registered_time)/1000)).strftime("%b %d, %Y")
        i = 0
        memberList = []
        memberIdList = []
        for member in members:
            i = i + 1
            if i <= 25:
                if member.cohort.lower() == cohort.lower():
                    memberList.append(member)
                    memberIdList.append(member.pk)
        r = int(len(memberList) / 25)
        m = len(memberList) % 25
        if m > 0:
            r = r + 2
        else:
            r = r + 1

        if len(memberList) == 0:
            return render(request, 'motherwise/result.html',
                          {'response': 'The cohort\'s members don\'t exist.'})

        request.session['selected_member_list'] = memberIdList
        request.session['selected_option'] = option
        if option == 'members':
            return render(request, 'motherwise/adminhome.html', {'me':admin,'users':memberList, 'range': range(r), 'current': 1, 'cohort': cohort})
        elif option == 'video':
            return redirect('/open_conference?group_id=0&cohort=' + cohort)
        elif option == 'group_chat':
            groups = Group.objects.filter(member_id=admin.pk).order_by('-id')
            for group in groups:
                group.created_time = datetime.datetime.fromtimestamp(float(int(group.created_time)/1000)).strftime("%b %d, %Y %H:%M")
                group.last_connected_time = datetime.datetime.fromtimestamp(float(int(group.last_connected_time)/1000)).strftime("%b %d, %Y %H:%M")
            latestGroupMemberList = []
            latest_group = None
            if groups.count() > 0:
                latest_group = groups[0]
                gMembers = GroupMember.objects.filter(group_id=latest_group.pk)
                for gMember in gMembers:
                    members = Member.objects.filter(id=gMember.member_id)
                    if members.count() > 0:
                        latestGroupMemberList.append(members[0])
            gcs = GroupConnect.objects.filter(member_id=admin.pk).order_by('-id')
            recents = []
            for gc in gcs:
                gs = Group.objects.filter(id=gc.group_id)
                if gs.count() > 0:
                    group = gs[0]
                    group.created_time = datetime.datetime.fromtimestamp(float(int(group.created_time)/1000)).strftime("%b %d, %Y %H:%M")
                    group.last_connected_time = datetime.datetime.fromtimestamp(float(int(group.last_connected_time)/1000)).strftime("%b %d, %Y %H:%M")
                    recents.append(group)
            return render(request, 'motherwise/groups.html', {'members':memberList, 'group':latest_group, 'groups': groups, 'group_members':latestGroupMemberList, 'recents':recents})
        elif option == 'private_chat':
            contacts = update_admin_contact(admin, "")
            return render(request, 'motherwise/chat.html', {'members':memberList, 'me': admin, 'friend':memberList[0], 'contacts':contacts})

    else:
        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        memberIdList = []
        try:
            memberIdList = request.session['selected_member_list']
        except KeyError:
            print('No key')

        selectedOption = request.session['selected_option']

        memberList = []
        for member_id in memberIdList:
            members = Member.objects.filter(id=member_id)
            if members.count() > 0:
                member = members[0]
                if member.registered_time != '':
                    member.registered_time = datetime.datetime.fromtimestamp(float(int(member.registered_time)/1000)).strftime("%b %d, %Y")
                memberList.append(member)

        contacts = update_admin_contact(admin, "")

        if len(memberList) > 0:
            if selectedOption == 'private_chat':
                return render(request, 'motherwise/chat.html', {'members':memberList, 'me': admin, 'friend':memberList[0], 'contacts':contacts})
            elif selectedOption == 'group_chat':
                groups = Group.objects.filter(member_id=admin.pk).order_by('-id')
                for group in groups:
                    group.created_time = datetime.datetime.fromtimestamp(float(int(group.created_time)/1000)).strftime("%b %d, %Y %H:%M")
                    group.last_connected_time = datetime.datetime.fromtimestamp(float(int(group.last_connected_time)/1000)).strftime("%b %d, %Y %H:%M")
                latestGroupMemberList = []
                latest_group = None
                if groups.count() > 0:
                    latest_group = groups[0]
                    gMembers = GroupMember.objects.filter(group_id=latest_group.pk)
                    for gMember in gMembers:
                        members = Member.objects.filter(id=gMember.member_id)
                        if members.count() > 0:
                            latestGroupMemberList.append(members[0])
                gcs = GroupConnect.objects.filter(member_id=admin.pk).order_by('-id')
                recents = []
                for gc in gcs:
                    gs = Group.objects.filter(id=gc.group_id)
                    if gs.count() > 0:
                        group = gs[0]
                        group.created_time = datetime.datetime.fromtimestamp(float(int(group.created_time)/1000)).strftime("%b %d, %Y %H:%M")
                        group.last_connected_time = datetime.datetime.fromtimestamp(float(int(group.last_connected_time)/1000)).strftime("%b %d, %Y %H:%M")
                        recents.append(group)
                return render(request, 'motherwise/groups.html', {'members':memberList, 'group':latest_group, 'groups': groups, 'group_members':latestGroupMemberList, 'recents':recents})
            elif selectedOption == 'members':
                r = int(len(memberList) / 25)
                m = len(memberList) % 25
                if m > 0:
                    r = r + 2
                else:
                    r = r + 1
                cohort = memberList[0].cohort
                return render(request, 'motherwise/adminhome.html', {'me':admin,'users':memberList, 'range': range(r), 'current': 1, 'cohort': cohort})
            elif selectedOption == 'video':
                cohort = memberList[0].cohort
                return redirect('/open_conference?group_id=0&cohort=' + cohort)
            else:
                return redirect('/home')
        else:
            return redirect('/home')



@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def search_members(request):
    if request.method == 'POST':
        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        search_id = request.POST.get('q', None)

        memberList = []

        members = Member.objects.filter(admin_id=adminID).order_by('-id')
        memberList = get_filtered_members_data(members, search_id)
        users, range = get_member_data(memberList)
        return render(request, 'motherwise/adminhome.html', {'me':admin,'users':users, 'range': range, 'current': 1, 'title':'Searched by ' + search_id})



def get_filtered_members_data(members, keyword):
    import datetime
    memberList = []
    for member in members:
        if keyword.lower() in member.name.lower():
            memberList.append(member)
        elif keyword.lower() in member.name.lower():
            memberList.append(member)
        elif keyword.lower() in member.email.lower():
            memberList.append(member)
        elif keyword.lower() in member.phone_number.lower():
            memberList.append(member)
        elif keyword.lower() in member.cohort.lower():
            memberList.append(member)
        elif keyword.lower() in member.address.lower():
            memberList.append(member)
        else:
            if member.registered_time != '':
                date_time = datetime.datetime.fromtimestamp(float(int(member.registered_time)/1000)).strftime("%b %d, %Y %H:%M")
                if keyword.lower() in date_time.lower():
                    memberList.append(member)

    return memberList


def get_member_data(memblist):
    import datetime
    i = 0
    memberList = []
    for member in memblist:
        if member.registered_time != '':
            member.registered_time = datetime.datetime.fromtimestamp(float(int(member.registered_time)/1000)).strftime("%b %d, %Y")
        i = i + 1
        if i <= 25:
            memberList.append(member)
    r = int(len(memblist) / 25)
    m = len(memblist) % 25
    if m > 0:
        r = r + 2
    else:
        r = r + 1

    return memberList, range(r)



def admin_account(request):
    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']

    admins = Member.objects.filter(id=adminID)
    if admins.count() > 0:
        admin = admins[0]
    return  render(request, 'motherwise/account.html', {'admin':admin})


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def edit_admin_account(request):
    if request.method == 'POST':
        email = request.POST.get('email', '')
        oldpassword = request.POST.get('oldpassword', '')
        newpassword = request.POST.get('newpassword', '')

        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']

        admin = Member.objects.get(id=adminID)
        if email == admin.email and oldpassword == admin.password:
            admin.password = newpassword

            admin.save()

        elif email == admin.email and oldpassword != admin.password:
            return render(request, 'motherwise/result.html',
                          {'response': 'Your old password is incorrect. Please enter your correct password.'})

        else:
            return render(request, 'motherwise/result.html',
                          {'response': 'Your email or password is incorrect. Please enter your correct information.'})

        return  render(request, 'motherwise/account.html', {'admin':admin, 'note':'account_updated'})


def torequestpwd(request):
    return  render(request, 'motherwise/forgot_password.html')


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def send_mail_forgotpassword(request):
    if request.method == 'POST':
        email = request.POST.get('email', '')

        members = Member.objects.filter(email=email)
        if members.count() == 0:
            return render(request, 'motherwise/result.html',
                          {'response': 'This email doesn\'t exist. Please try another one.'})

        message = 'You are allowed to reset your password from your request.<br>For it, please click this link to reset your password.<br><br><a href=\'' + 'https://www.vacaydemo.com/resetpassword?email=' + email
        message = message + '\' target=\'_blank\'>' + 'Link to reset password' + '</a>'

        html =  """\
                    <html>
                        <head></head>
                        <body>
                            <a href="#"><img src="https://www.vacaydemo.com/static/images/vacaylogo.jpg" style="width:120px;height:120px; margin-left:25px; border-radius:8%;"/></a>
                            <h2 style="color:#02839a;">VaCay Administrator's Security Update Information</h2>
                            <div style="font-size:14px; white-space: pre-line; word-wrap: break-word;">
                                {mes}
                            </div>
                        </body>
                    </html>
                """
        html = html.format(mes=message)

        fromEmail = settings.ADMIN_EMAIL
        toEmailList = []
        toEmailList.append(email)
        msg = EmailMultiAlternatives('We allowed you to reset your password', '', fromEmail, toEmailList)
        msg.attach_alternative(html, "text/html")
        msg.send(fail_silently=False)

        return render(request, 'motherwise/result.html',
                          {'response': 'We sent a message to your email. Please check and reset your password.'})


def resetpassword(request):
    email = request.GET['email']
    return render(request, 'motherwise/resetpwd.html', {'email':email})


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def admin_rstpwd(request):
    if request.method == 'POST':
        email = request.POST.get('email', '')
        password = request.POST.get('password', '')

        members = Member.objects.filter(email=email)
        if members.count() == 0:
            return redirect('/signuppage/')

        member = members[0]
        member.password = password
        member.save()

        return render(request, 'motherwise/admin.html', {'notify':'password changed'})



@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def send_cohort_message(request):

    import datetime

    if request.method == 'POST':

        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        cohort = request.POST.get('cohort', '')
        message = request.POST.get('message', '')

        members = Member.objects.filter(cohort=cohort)

        for member in members:

            notification = Notification()
            notification.member_id = member.pk
            notification.sender_id = admin.pk
            notification.message = message
            notification.notified_time = str(int(round(time.time() * 1000)))
            notification.save()

            rcv = Received()
            rcv.member_id = member.pk
            rcv.sender_id = admin.pk
            rcv.noti_id = notification.pk
            rcv.save()

            snt = Sent()
            snt.member_id = member.pk
            snt.sender_id = admin.pk
            snt.noti_id = notification.pk
            snt.save()

            title = 'You\'ve received a message from VaCay Community'
            subject = 'From VaCay Community'
            msg = 'Dear ' + member.name + ',<br><br>'
            msg = msg + message
            msg = msg + '<br><a href=\'' + settings.URL + '/users/notifications?noti_id=' + str(notification.pk) + '\' target=\'_blank\'>Join website</a>' + '<br><br>VaCay Community'

            from_email = admin.email
            to_emails = []
            to_emails.append(member.email)
            send_mail_message(from_email, to_emails, title, subject, msg)

            msg = member.name + ', You\'ve received a message from VaCay Community.\nThe message is as following:\n' + message

            ##########################################################################################################################################################################

            db = firebase.database()
            data = {
                "msg": msg,
                "date":str(int(round(time.time() * 1000))),
                "sender_id": str(admin.pk),
                "sender_name": admin.name,
                "sender_email": admin.email,
                "sender_photo": admin.photo_url,
                "role": "admin",
                "type": "message",
                "id": str(notification.pk),
                "mes_id": str(notification.pk)
            }

            db.child("notify").child(str(member.pk)).push(data)
            db.child("notify2").child(str(member.pk)).push(data)

            sendFCMPushNotification(member.pk, admin.pk, msg)

            #################################################################################################################################################################################

            if member.playerID != '':
                playerIDList = []
                playerIDList.append(member.playerID)
                url = '/users/notifications?noti_id=' + str(notification.pk)
                send_push(playerIDList, msg, url)

        members = Member.objects.filter(admin_id=adminID).order_by('-id')
        for member in members:
            if member.registered_time != '':
                member.registered_time = datetime.datetime.fromtimestamp(float(int(member.registered_time)/1000)).strftime("%b %d, %Y")
        users, range = get_all_member_data(members)

        return render(request, 'motherwise/adminhome.html', {'me':admin,'users':users, 'range': range, 'current': 1, 'notify':'message_sent'})


def admin_switch_chat(request):

    member_id = request.GET['member_id']

    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    members = Member.objects.filter(id=member_id)
    if members.count() == 0:
        return redirect('/home')

    selected_member = members[0]

    memberIdList = []
    try:
        memberIdList = request.session['selected_member_list']
    except KeyError:
        print('No key')

    selectedOption = request.session['selected_option']

    if len(memberIdList) == 0:
        return redirect('/home')

    memberList = []
    for member_id in memberIdList:
        members = Member.objects.filter(id=member_id)
        if members.count() > 0:
            member = members[0]
            memberList.append(member)

    for member in memberList:
        if selected_member.pk == member.pk:
            index = memberList.index(member)
            del memberList[index]
            memberList.insert(0, selected_member)

    if selected_member not in memberList:
        memberList.insert(0,selected_member)
        memberIdList.insert(0, selected_member.pk)

    contacts = update_admin_contact(admin, "")

    if len(memberList) == 0:
        return render(request, 'motherwise/result.html', {'response': 'The member doesn\'t exist.'})

    if selectedOption == 'private_chat':
        return render(request, 'motherwise/chat.html', {'members':memberList, 'me': admin, 'friend':memberList[0], 'contacts':contacts})
    else:
        return redirect('/home')


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def admin_to_chat(request):
    if request.method == 'POST':

        email = request.POST.get('member_email', '')

        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        members = Member.objects.filter(email=email)
        if members.count() == 0:
            return redirect('/home')

        member = members[0]
        contacts = update_admin_contact(admin, email)

        memberList = []
        memberList.insert(0, member)

        return render(request, 'motherwise/chat.html', {'members':memberList, 'me': admin, 'friend':memberList[0], 'contacts':contacts})

    else:
        return redirect('/home')



def update_admin_contact(admin, member_email):
    if member_email != '':
        contacts = Contact.objects.filter(member_id=admin.pk, contact_email=member_email)
        if contacts.count() == 0:
            contact = Contact()
            contact.member_id = admin.pk
            contact.contact_email = member_email
            contact.contacted_time = str(int(round(time.time() * 1000)))
            contact.save()
        else:
            contact = contacts[0]
            contacts = Contact.objects.filter(member_id=admin.pk)
            recent_contact = contacts[contacts.count() - 1]
            if contact.pk < recent_contact.pk:
                contact.delete()
                contact = Contact()
                contact.member_id = admin.pk
                contact.contact_email = member_email
                contact.contacted_time = str(int(round(time.time() * 1000)))
                contact.save()

    contacts = Contact.objects.filter(member_id=admin.pk).order_by('-id')
    contactList = []
    for contact in contacts:
        members = Member.objects.filter(email=contact.contact_email)
        if members.count() > 0:
            member = members[0]
            contactList.append(member)

    return contactList



def admin_switch_to_cohort(request):

    cohort = request.GET['cohort']

    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    members = Member.objects.filter(admin_id=adminID).order_by('-id')
    memberList = []
    memberIdList = []
    for member in members:
        if member.cohort.lower() == cohort.lower():
            memberList.append(member)
            memberIdList.append(member.pk)

    if len(memberList) == 0:
            return render(request, 'motherwise/result.html',
                          {'response': 'The cohort\'s members don\'t exist.'})

    request.session['selected_member_list'] = memberIdList

    contacts = update_admin_contact(admin, "")

    return render(request, 'motherwise/chat.html', {'members':memberList, 'me': admin, 'friend':memberList[0], 'contacts':contacts})



@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def create_group(request):

    import datetime

    if request.method == 'POST':

        name = request.POST.get('name', '')

        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        groups = Group.objects.filter(member_id=admin.pk, name=name)
        if groups.count() > 0:
            return render(request, 'motherwise/result.html',
                          {'response': 'The same name already exists.'})

        memberIdList = []
        try:
            memberIdList = request.session['selected_member_list']
        except KeyError:
            print('No key')

        memberList = []
        for member_id in memberIdList:
            members = Member.objects.filter(id=member_id)
            if members.count() > 0:
                member = members[0]
                memberList.append(member)

        group = Group()
        group.member_id = admin.pk
        group.name = name
        group.code = get_group_code(name)
        group.color = get_group_color()
        group.member_count = '0'
        group.created_time = str(int(round(time.time() * 1000)))
        group.last_connected_time = str(int(round(time.time() * 1000)))
        group.save()

        groups = Group.objects.filter(member_id=admin.pk).order_by('-id')
        for group in groups:
            group.created_time = datetime.datetime.fromtimestamp(float(int(group.created_time)/1000)).strftime("%b %d, %Y %H:%M")
            group.last_connected_time = datetime.datetime.fromtimestamp(float(int(group.last_connected_time)/1000)).strftime("%b %d, %Y %H:%M")
        latestGroupMemberList = []
        latest_group = None
        if groups.count() > 0:
            latest_group = groups[0]
            gMembers = GroupMember.objects.filter(group_id=latest_group.pk)
            for gMember in gMembers:
                members = Member.objects.filter(id=gMember.member_id)
                if members.count() > 0:
                    latestGroupMemberList.append(members[0])
        gcs = GroupConnect.objects.filter(member_id=admin.pk).order_by('-id')
        recents = []
        for gc in gcs:
            gs = Group.objects.filter(id=gc.group_id)
            if gs.count() > 0:
                g = gs[0]
                g.created_time = datetime.datetime.fromtimestamp(float(int(g.created_time)/1000)).strftime("%b %d, %Y %H:%M")
                g.last_connected_time = datetime.datetime.fromtimestamp(float(int(g.last_connected_time)/1000)).strftime("%b %d, %Y %H:%M")
                recents.append(g)
        return render(request, 'motherwise/groups.html', {'members':memberList, 'group':group, 'groups': groups, 'group_members':latestGroupMemberList, 'recents':recents})

    else:
        return redirect('/home')


def get_group_code(name):
    name = name.split(' ')
    if len(name) > 1:
        name = name[0][0:1] + name[1][0:1]
    else: name = name[0][0:1]
    name = name.upper()
    return name

def get_group_color():
    from random import randint
    color = '#{:06x}'.format(randint(0, 256**3))
    return color


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def message_to_group(request):

    import datetime

    if request.method == 'POST':

        group_id = request.POST.get('group_id', '1')
        message = request.POST.get('message', '')
        option = request.POST.get('option', '')

        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        groups = Group.objects.filter(id=group_id)
        if groups.count() > 0:
            group = groups[0]

            if option == 'invite':

                memberIdList = []
                try:
                    memberIdList = request.session['selected_member_list']
                except KeyError:
                    print('No key')

                memberList = []
                for member_id in memberIdList:
                    members = Member.objects.filter(id=member_id)
                    if members.count() > 0:
                        member = members[0]
                        gMembers = GroupMember.objects.filter(group_id=group.pk, member_id=member.pk)
                        if gMembers.count() == 0:
                            gMember = GroupMember()
                            gMember.group_id = group.pk
                            gMember.member_id = member.pk
                            gMember.invited_time = str(int(round(time.time() * 1000)))
                            gMember.save()

                            notification = Notification()
                            notification.member_id = member.pk
                            notification.sender_id = admin.pk
                            notification.notified_time = str(int(round(time.time() * 1000)))
                            notification.save()

                            rcv = Received()
                            rcv.member_id = member.pk
                            rcv.sender_id = admin.pk
                            rcv.noti_id = notification.pk
                            rcv.save()

                            snt = Sent()
                            snt.member_id = member.pk
                            snt.sender_id = admin.pk
                            snt.noti_id = notification.pk
                            snt.save()

                            title = 'You\'ve received an invitation from VaCay Community'
                            subject = 'From VaCay Community'
                            msg = 'Dear ' + member.name + ',<br><br>You\'ve received an invitation to: ' + group.name + ' from VaCay Community.<br><br>Community name: ' + group.name + '<br><br>'
                            # msg = msg + 'So you can see that community in your account and connect it to became a member of the community, attending all the events from it.<br>'
                            msg = msg + 'There’s nothing you have to do- you’re already included. The next time you login, just click on the \"Communities\" icon and you’ll see this option.<br>'
                            msg = msg + 'From there, you’ll see the members, see which videos have been posted, engage in a group chat and send private messages.<br>'
                            msg = msg + 'These communities are ways you can share similar interests and get to know your VaCay Family. Have fun!<br>'
                            msg = msg + '<a href=\'' + settings.URL + '/users/notifications?noti_id=' + str(notification.pk) + '\' target=\'_blank\'>Join website</a>' + '<br><br>VaCay Team'

                            title2 = 'Recibiste una invitación de VaCay Community'
                            msg2 = 'Querida ' + member.name + ',<br><br>Recibió una invitación para una comunidad ' + group.name + ' de la comunidad VaCay.<br><br>Nombre de la comunidad: ' + group.name + '<br><br>'
                            # msg2 = msg2 + 'No hay nada que tenga que hacer, ya está incluida. La próxima vez que inicie sesión, simplemente haga clic en el icono de la comunidad y verá esta opción.<br>'
                            msg2 = msg2 + 'No hay nada que tengas que hacer, ya estás incluido. La próxima vez que inicie sesión, simplemente haga clic en el icono \"Comunidades\" y verá esta opción.<br>'
                            msg2 = msg2 + 'A partir de ahí, verá los miembros, verá qué videos se han publicado, participará en un chat grupal y enviará mensajes privados.<br>'
                            msg2 = msg2 + 'Estas comunidades son formas en que puede compartir intereses similares y conocer a su familia VaCay. ¡Que te diviertas!<br>'
                            msg2 = msg2 + '<a href=\'' + settings.URL + '/users/notifications?noti_id=' + str(notification.pk) + '\' target=\'_blank\'>Unirse al sitio web</a>' + '<br><br>VaCay Team'

                            from_email = admin.email
                            to_emails = []
                            to_emails.append(member.email)
                            # send_mail_message(from_email, to_emails, title, subject, msg)
                            send_mail_message0(from_email, to_emails, title, subject, msg, title2, msg2)

                            msg = 'Dear ' + member.name + ',\n\nYou\'ve received an invitation to: ' + group.name + ' from VaCay Community.\n\nCommunity name: ' + group.name + '\n\n'
                            msg = msg + 'So you can see that community in your account and connect it to became a member of the community, attending all the events from it.\n'
                            msg = msg + 'There’s nothing you have to do- you’re already included. The next time you login, just click on the \"Communities\" icon and you’ll see this option.\n'
                            msg = msg + 'From there, you’ll see the members, see which videos have been posted, engage in a group chat and send private messages.\n'
                            msg = msg + 'These communities are ways you can share similar interests and get to know your VaCay Family. Have fun!\n'
                            msg = msg + 'Click on this link to join: ' + settings.URL + '/users/notifications?noti_id=' + str(notification.pk) + '\n\nVaCay Team'

                            notification.message = msg
                            notification.save()

                            ##########################################################################################################################################################################

                            db = firebase.database()
                            data = {
                                "msg": msg,
                                "date":str(int(round(time.time() * 1000))),
                                "sender_id": str(admin.pk),
                                "sender_name": admin.name,
                                "sender_email": admin.email,
                                "sender_photo": admin.photo_url,
                                "role": "admin",
                                "type": "group_invite",
                                "id": str(group.pk),
                                "mes_id": str(notification.pk)
                            }

                            db.child("notify").child(str(member.pk)).push(data)
                            db.child("notify2").child(str(member.pk)).push(data)

                            sendFCMPushNotification(member.pk, admin.pk, msg)

                            #################################################################################################################################################################################

                            if member.playerID != '':
                                playerIDList = []
                                playerIDList.append(member.playerID)
                                url = '/users/notifications?noti_id=' + str(notification.pk)
                                send_push(playerIDList, msg, url)

                        gms = GroupMember.objects.filter(group_id=group.pk, member_id=member.pk)
                        if gms.count() == 0: memberList.append(member)

                gMembers = GroupMember.objects.filter(group_id=group.pk).order_by('-id')
                group.member_count = str(gMembers.count())
                group.last_connected_time = str(int(round(time.time() * 1000)))
                group.save()

                update_recent_group(admin, group)

                latestGroupMemberList = []
                groups = Group.objects.filter(member_id=admin.pk).order_by('-id')
                for gr in groups:
                    gr.created_time = datetime.datetime.fromtimestamp(float(int(gr.created_time)/1000)).strftime("%b %d, %Y %H:%M")
                    gr.last_connected_time = datetime.datetime.fromtimestamp(float(int(gr.last_connected_time)/1000)).strftime("%b %d, %Y %H:%M")
                for gMember in gMembers:
                    members = Member.objects.filter(id=gMember.member_id)
                    if members.count() > 0:
                        latestGroupMemberList.append(members[0])

                gcs = GroupConnect.objects.filter(member_id=admin.pk).order_by('-id')
                recents = []
                for gc in gcs:
                    gs = Group.objects.filter(id=gc.group_id)
                    if gs.count() > 0:
                        g = gs[0]
                        g.created_time = datetime.datetime.fromtimestamp(float(int(g.created_time)/1000)).strftime("%b %d, %Y %H:%M")
                        g.last_connected_time = datetime.datetime.fromtimestamp(float(int(g.last_connected_time)/1000)).strftime("%b %d, %Y %H:%M")
                        recents.append(g)

                return render(request, 'motherwise/groups.html', {'members':[], 'group':group, 'groups': groups, 'group_members':latestGroupMemberList, 'recents':recents})

            elif option == 'group':
                gMembers = GroupMember.objects.filter(group_id=group.pk).order_by('-id')
                for gm in gMembers:
                    members = Member.objects.filter(id=gm.member_id)
                    if members.count() > 0:
                        member = members[0]

                        notification = Notification()
                        notification.member_id = member.pk
                        notification.sender_id = admin.pk
                        notification.message = message
                        notification.notified_time = str(int(round(time.time() * 1000)))
                        notification.save()

                        rcv = Received()
                        rcv.member_id = member.pk
                        rcv.sender_id = admin.pk
                        rcv.noti_id = notification.pk
                        rcv.save()

                        snt = Sent()
                        snt.member_id = member.pk
                        snt.sender_id = admin.pk
                        snt.noti_id = notification.pk
                        snt.save()

                        title = 'You\'ve received a message from VaCay Community'
                        subject = 'From VaCay Community'
                        msg = 'Dear ' + member.name + ',<br><br>'
                        msg = msg + message
                        msg = msg + '<br><a href=\'' + settings.URL + '/users/notifications?noti_id=' + str(notification.pk) + '\' target=\'_blank\'>Join website</a>' + '<br><br>VaCay Team'

                        from_email = admin.email
                        to_emails = []
                        to_emails.append(member.email)
                        send_mail_message(from_email, to_emails, title, subject, msg)

                        msg = member.name + ', You\'ve received a message from VaCay Community.\nThe message is as following:\n' + message

                        ##########################################################################################################################################################################

                        db = firebase.database()
                        data = {
                            "msg": message,
                            "date":str(int(round(time.time() * 1000))),
                            "sender_id": str(admin.pk),
                            "sender_name": admin.name,
                            "sender_email": admin.email,
                            "sender_photo": admin.photo_url,
                            "role": "admin",
                            "type": "message",
                            "id": str(notification.pk),
                            "mes_id": str(notification.pk)
                        }

                        db.child("notify").child(str(member.pk)).push(data)
                        db.child("notify2").child(str(member.pk)).push(data)

                        sendFCMPushNotification(member.pk, admin.pk, message)

                        #################################################################################################################################################################################

                        if member.playerID != '':
                            playerIDList = []
                            playerIDList.append(member.playerID)
                            url = '/users/notifications?noti_id=' + str(notification.pk)
                            send_push(playerIDList, msg, url)

                update_recent_group(admin, group)

                memberIdList = []
                try:
                    memberIdList = request.session['selected_member_list']
                except KeyError:
                    print('No key')

                memberList = []
                for member_id in memberIdList:
                    members = Member.objects.filter(id=member_id)
                    if members.count() > 0:
                        member = members[0]
                        gms = GroupMember.objects.filter(group_id=group.pk, member_id=member.pk)
                        if gms.count() == 0: memberList.append(member)

                latestGroupMemberList = []
                groups = Group.objects.filter(member_id=admin.pk).order_by('-id')
                for gr in groups:
                    gr.created_time = datetime.datetime.fromtimestamp(float(int(gr.created_time)/1000)).strftime("%b %d, %Y %H:%M")
                    gr.last_connected_time = datetime.datetime.fromtimestamp(float(int(gr.last_connected_time)/1000)).strftime("%b %d, %Y %H:%M")
                for gMember in gMembers:
                    members = Member.objects.filter(id=gMember.member_id)
                    if members.count() > 0:
                        latestGroupMemberList.append(members[0])

                gcs = GroupConnect.objects.filter(member_id=admin.pk).order_by('-id')
                recents = []
                for gc in gcs:
                    gs = Group.objects.filter(id=gc.group_id)
                    if gs.count() > 0:
                        g = gs[0]
                        g.created_time = datetime.datetime.fromtimestamp(float(int(g.created_time)/1000)).strftime("%b %d, %Y %H:%M")
                        g.last_connected_time = datetime.datetime.fromtimestamp(float(int(g.last_connected_time)/1000)).strftime("%b %d, %Y %H:%M")
                        recents.append(g)

                return render(request, 'motherwise/groups.html', {'members':memberList, 'group':group, 'groups': groups, 'group_members':latestGroupMemberList, 'recents':recents})

            else: return redirect('/home')

        else:
            return redirect('/home')

    else:
        return redirect('/home')


def update_recent_group(admin, group):
    gConnects = GroupConnect.objects.filter(member_id=admin.pk, group_id=group.pk)
    if gConnects.count() > 0:
        gConnect = gConnects[0]
        recent = gConnects[gConnects.count() - 1]
        if gConnect.pk < recent.pk:
            gConnect.delete()
            gc = GroupConnect()
            gc.member_id = admin.pk
            gc.group_id = group.pk
            gc.last_connected_time = str(int(round(time.time() * 1000)))
            gc.save()
    else:
        gc = GroupConnect()
        gc.member_id = admin.pk
        gc.group_id = group.pk
        gc.last_connected_time = str(int(round(time.time() * 1000)))
        gc.save()



def switch_group(request):

    import datetime

    group_id = request.GET['group_id']

    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    groups = Group.objects.filter(id=group_id)
    if groups.count() == 0:
        return redirect('/home')

    selected_group = groups[0]
    selected_group.created_time = datetime.datetime.fromtimestamp(float(int(selected_group.created_time)/1000)).strftime("%b %d, %Y %H:%M")
    selected_group.last_connected_time = datetime.datetime.fromtimestamp(float(int(selected_group.last_connected_time)/1000)).strftime("%b %d, %Y %H:%M")

    memberIdList = []
    try:
        memberIdList = request.session['selected_member_list']
    except KeyError:
        print('No key')

    selectedOption = request.session['selected_option']

    # if len(memberIdList) == 0:
    #     return redirect('/home')

    memberList = []
    for member_id in memberIdList:
        members = Member.objects.filter(id=member_id)
        if members.count() > 0:
            member = members[0]
            gms = GroupMember.objects.filter(group_id=selected_group.pk, member_id=member.pk)
            if gms.count() == 0: memberList.append(member)

    groups = Group.objects.filter(member_id=admin.pk).order_by('-id')
    groupList = []
    for group in groups:
        group.created_time = datetime.datetime.fromtimestamp(float(int(group.created_time)/1000)).strftime("%b %d, %Y %H:%M")
        group.last_connected_time = datetime.datetime.fromtimestamp(float(int(group.last_connected_time)/1000)).strftime("%b %d, %Y %H:%M")
        groupList.append(group)
    for group in groupList:
        if selected_group.pk == group.pk:
            index = groupList.index(group)
            del groupList[index]
            groupList.insert(0, selected_group)

    latestGroupMemberList = []
    gMembers = GroupMember.objects.filter(group_id=selected_group.pk)
    for gMember in gMembers:
        members = Member.objects.filter(id=gMember.member_id)
        if members.count() > 0:
            latestGroupMemberList.append(members[0])
    gcs = GroupConnect.objects.filter(member_id=admin.pk).order_by('-id')
    recents = []
    for gc in gcs:
        gs = Group.objects.filter(id=gc.group_id)
        if gs.count() > 0:
            group = gs[0]
            group.created_time = datetime.datetime.fromtimestamp(float(int(group.created_time)/1000)).strftime("%b %d, %Y %H:%M")
            group.last_connected_time = datetime.datetime.fromtimestamp(float(int(group.last_connected_time)/1000)).strftime("%b %d, %Y %H:%M")
            recents.append(group)
    return render(request, 'motherwise/groups.html', {'members':memberList, 'group':selected_group, 'groups': groupList, 'group_members':latestGroupMemberList, 'recents':recents})


def get_groups(request):
    import datetime
    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    groups = Group.objects.filter(member_id=admin.pk).order_by('-id')
    for group in groups:
        group.created_time = datetime.datetime.fromtimestamp(float(int(group.created_time)/1000)).strftime("%b %d, %Y %H:%M")
        group.last_connected_time = datetime.datetime.fromtimestamp(float(int(group.last_connected_time)/1000)).strftime("%b %d, %Y %H:%M")
    latestGroupMemberList = []
    latest_group = None
    if groups.count() > 0:
        latest_group = groups[0]
        gMembers = GroupMember.objects.filter(group_id=latest_group.pk)
        for gMember in gMembers:
            members = Member.objects.filter(id=gMember.member_id)
            if members.count() > 0:
                latestGroupMemberList.append(members[0])
    gcs = GroupConnect.objects.filter(member_id=admin.pk).order_by('-id')
    recents = []
    for gc in gcs:
        gs = Group.objects.filter(id=gc.group_id)
        if gs.count() > 0:
            group = gs[0]
            group.created_time = datetime.datetime.fromtimestamp(float(int(group.created_time)/1000)).strftime("%b %d, %Y %H:%M")
            group.last_connected_time = datetime.datetime.fromtimestamp(float(int(group.last_connected_time)/1000)).strftime("%b %d, %Y %H:%M")
            recents.append(group)
    request.session['selected_option'] = 'group_chat'
    return render(request, 'motherwise/groups.html', {'members':[], 'group':latest_group, 'groups': groups, 'group_members':latestGroupMemberList, 'recents':recents})



def open_group_chat(request):
    group_id = request.GET['group_id']
    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    memberList = []
    group = None
    groups = Group.objects.filter(member_id=admin.pk, id=group_id).order_by('-id')
    if groups.count() > 0:
        group = groups[0]
        gMembers = GroupMember.objects.filter(group_id=group.pk)
        for gMember in gMembers:
            members = Member.objects.filter(id=gMember.member_id)
            if members.count() > 0:
                memb = members[0]
                memb.username = '@' + memb.email[0:memb.email.find('@')]
                memberList.append(memb)

        request.session['group_id'] = group_id
        request.session['cohort'] = ''

        memberIdList = []
        for memb in memberList:
            memberIdList.append(memb.pk)
        request.session['selected_member_list'] = memberIdList

        return render(request, 'motherwise/group_chat.html', {'me':admin, 'members':memberList, 'group':group})
    else:
        return redirect('/home')


def group_cohort_chat(request):
    cohort = request.GET['cohort']
    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    members = Member.objects.filter(admin_id=admin.pk, cohort=cohort).order_by('-id')
    for member in members:
        member.username = '@' + member.email[0:member.email.find('@')]

    request.session['cohort'] = cohort
    request.session['group_id'] = ''

    memberIdList = []
    for memb in members:
        memberIdList.append(memb.pk)
    request.session['selected_member_list'] = memberIdList

    return render(request, 'motherwise/group_chat.html', {'me':admin, 'members':members, 'cohort':cohort})



@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def group_chat_message(request):
    if request.method == 'POST':

        group_id = request.POST.get('group_id', '')
        message = request.POST.get('message', '')
        option = request.POST.get('option', '')

        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        if option == 'group':
            groups = Group.objects.filter(id=group_id)
            if groups.count() > 0:
                group = groups[0]
                gMembers = GroupMember.objects.filter(group_id=group.pk)
                for gMember in gMembers:
                    members = Member.objects.filter(id=gMember.member_id)
                    if members.count() > 0:
                        member = members[0]

                        notification = Notification()
                        notification.member_id = member.pk
                        notification.sender_id = admin.pk
                        notification.message = message
                        notification.notified_time = str(int(round(time.time() * 1000)))
                        notification.save()

                        rcv = Received()
                        rcv.member_id = member.pk
                        rcv.sender_id = admin.pk
                        rcv.noti_id = notification.pk
                        rcv.save()

                        snt = Sent()
                        snt.member_id = member.pk
                        snt.sender_id = admin.pk
                        snt.noti_id = notification.pk
                        snt.save()

                        title = 'You\'ve received a message from VaCay Community'
                        subject = 'From VaCay Community'
                        msg = 'Dear ' + member.name + ',<br><br>'
                        msg = msg + message
                        msg = msg + '<br><a href=\'' + settings.URL + '/users/notifications?noti_id=' + str(notification.pk) + '\' target=\'_blank\'>Join website</a>' + '<br><br>VaCay Team'

                        from_email = admin.email
                        to_emails = []
                        to_emails.append(member.email)
                        send_mail_message(from_email, to_emails, title, subject, msg)

                        msg = member.name + ', You\'ve received a message from VaCay Community.\nThe message is as following:\n' + message

                        ##########################################################################################################################################################################

                        db = firebase.database()
                        data = {
                            "msg": message,
                            "date":str(int(round(time.time() * 1000))),
                            "sender_id": str(admin.pk),
                            "sender_name": admin.name,
                            "sender_email": admin.email,
                            "sender_photo": admin.photo_url,
                            "role": "admin",
                            "type": "message",
                            "id": str(notification.pk),
                            "mes_id": str(notification.pk)
                        }

                        db.child("notify").child(str(member.pk)).push(data)
                        db.child("notify2").child(str(member.pk)).push(data)

                        sendFCMPushNotification(member.pk, admin.pk, message)

                        #################################################################################################################################################################################

                        if member.playerID != '':
                            playerIDList = []
                            playerIDList.append(member.playerID)
                            url = '/users/notifications?noti_id=' + str(notification.pk)
                            send_push(playerIDList, msg, url)

            else:
                return redirect('/home')

        elif option == 'cohort':
            members = Member.objects.filter(cohort=group_id)
            for member in members:

                notification = Notification()
                notification.member_id = member.pk
                notification.sender_id = admin.pk
                notification.message = message
                notification.notified_time = str(int(round(time.time() * 1000)))
                notification.save()

                rcv = Received()
                rcv.member_id = member.pk
                rcv.sender_id = admin.pk
                rcv.noti_id = notification.pk
                rcv.save()

                snt = Sent()
                snt.member_id = member.pk
                snt.sender_id = admin.pk
                snt.noti_id = notification.pk
                snt.save()

                title = 'You\'ve received a message from VaCay Community'
                subject = 'From VaCay Community'
                msg = 'Dear ' + member.name + ',<br><br>'
                msg = msg + message
                msg = msg + '<br><a href=\'' + settings.URL + '/users/notifications?noti_id=' + str(notification.pk) + '\' target=\'_blank\'>Join website</a>' + '<br><br>VaCay Team'

                from_email = admin.email
                to_emails = []
                to_emails.append(member.email)
                send_mail_message(from_email, to_emails, title, subject, msg)

                msg = member.name + ', You\'ve received a message from VaCay Community.\nThe message is as following:\n' + message

                ##########################################################################################################################################################################

                db = firebase.database()
                data = {
                    "msg": message,
                    "date":str(int(round(time.time() * 1000))),
                    "sender_id": str(admin.pk),
                    "sender_name": admin.name,
                    "sender_email": admin.email,
                    "sender_photo": admin.photo_url,
                    "role": "admin",
                    "type": "message",
                    "id": str(notification.pk),
                    "mes_id": str(notification.pk)
                }

                db.child("notify").child(str(member.pk)).push(data)
                db.child("notify2").child(str(member.pk)).push(data)

                sendFCMPushNotification(member.pk, admin.pk, message)

                #################################################################################################################################################################################

                if member.playerID != '':
                    playerIDList = []
                    playerIDList.append(member.playerID)
                    url = '/users/notifications?noti_id=' + str(notification.pk)
                    send_push(playerIDList, msg, url)

        cohort = request.session['cohort']
        group_id = request.session['group_id']

        if cohort != '':
            return redirect('/group_cohort_chat?cohort=' + cohort)
        elif group_id is not None and int(group_id) > 0:
            return redirect('/open_group_chat?group_id=' + group_id)
        else:
            return redirect('/home')


def group_private_chat(request):

    email = request.GET['email']

    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    members = Member.objects.filter(email=email)
    if members.count() == 0:
        return redirect('/home')

    member = members[0]
    contacts = update_admin_contact(admin, email)

    memberList = []
    memberList.insert(0, member)

    memberIdList = []
    memberIdList.insert(0, member.pk)

    request.session['selected_option'] = 'private_chat'
    request.session['selected_member_list'] = memberIdList

    return render(request, 'motherwise/chat.html', {'members':memberList, 'me': admin, 'friend':memberList[0], 'contacts':contacts})



def admin_delete_contact(request):

    member_id = request.GET['member_id']

    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    members = Member.objects.filter(id=member_id)
    if members.count() > 0:
        member = members[0]
        contacts = Contact.objects.filter(member_id=admin.pk, contact_email=member.email)
        for contact in contacts:
            contact.delete()

    memberIdList = []
    try:
        memberIdList = request.session['selected_member_list']
    except KeyError:
        print('No key')

    memberList = []
    for member_id in memberIdList:
        members = Member.objects.filter(id=member_id)
        if members.count() > 0:
            member = members[0]
            memberList.append(member)

    contacts = update_admin_contact(admin, "")

    return render(request, 'motherwise/chat.html', {'members':memberList, 'me': admin, 'friend':memberList[0], 'contacts':contacts})


def admin_delete_group(request):

    group_id = request.GET['group_id']

    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    groups = Group.objects.filter(id=group_id)
    if groups.count() > 0:
        group = groups[0]
        gms = GroupMember.objects.filter(group_id=group.pk)
        for gm in gms:
            gm.delete()

        gcs = GroupConnect.objects.filter(group_id=group.pk)
        for gc in gcs:
            gc.delete()

        group.delete()

    return redirect('/groups/')



def to_posts(request):
    import datetime
    try:
        if request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    users = Member.objects.filter(admin_id=admin.pk).order_by('-id')
    userList = []
    for user in users:
        if user.registered_time != '':
            user.username = '@' + user.email[0:user.email.find('@')]
            userList.append(user)

    list1 = []
    list2 = []
    list3 = []
    list4 = []

    allPosts = Post.objects.all().order_by('-id')
    i = 0
    for post in allPosts:
        post.posted_time = datetime.datetime.fromtimestamp(float(int(post.posted_time)/1000)).strftime("%b %d, %Y %H:%M")
        i = i + 1
        pls = PostLike.objects.filter(post_id=post.pk, member_id=admin.pk)
        if pls.count() > 0:
            post.liked = 'yes'
        else: post.liked = 'no'

        comments = Comment.objects.filter(post_id=post.pk)
        post.comments = str(comments.count())
        likes = PostLike.objects.filter(post_id=post.pk)
        post.likes = str(likes.count())

        members = Member.objects.filter(id=post.member_id)
        if members.count() > 0:
            memb = members[0]
            if int(memb.admin_id) == admin.pk or memb.pk == admin.pk:
                data = {
                    'member':memb,
                    'post': post
                }

                if i % 4 == 1: list1.append(data)
                elif i % 4 == 2: list2.append(data)
                elif i % 4 == 3: list3.append(data)
                elif i % 4 == 0: list4.append(data)

    pst = None
    try:
        post_id = request.GET['post_id']
        posts = Post.objects.filter(id=post_id)
        if posts.count() > 0:
            pst = posts[0]
            pst.posted_time = datetime.datetime.fromtimestamp(float(int(pst.posted_time)/1000)).strftime("%b %d, %Y %H:%M")
    except KeyError:
        print('no key')

    return render(request, 'motherwise/post.html', {'me':admin, 'list1':list1, 'list2':list2, 'list3':list3, 'list4':list4, 'users':userList, 'pst':pst})



def my_posts(request):

    import datetime

    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    users = Member.objects.filter(admin_id=admin.pk).order_by('-id')
    userList = []
    for user in users:
        if user.registered_time != '':
            user.username = '@' + user.email[0:user.email.find('@')]
            userList.append(user)

    list1 = []
    list2 = []
    list3 = []
    list4 = []

    posts = Post.objects.filter(member_id=admin.pk).order_by('-id')

    i = 0
    for post in posts:
        post.posted_time = datetime.datetime.fromtimestamp(float(int(post.posted_time)/1000)).strftime("%b %d, %Y %H:%M")

        comments = Comment.objects.filter(post_id=post.pk)
        post.comments = str(comments.count())
        likes = PostLike.objects.filter(post_id=post.pk)
        post.likes = str(likes.count())

        i = i + 1

        data = {
            'member':admin,
            'post': post
        }

        if i % 4 == 1: list1.append(data)
        elif i % 4 == 2: list2.append(data)
        elif i % 4 == 3: list3.append(data)
        elif i % 4 == 0: list4.append(data)

    return render(request, 'motherwise/post.html', {'me':admin, 'list1':list1, 'list2':list2, 'list3':list3, 'list4':list4, 'search':'My', 'users':userList})



@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def create_post(request):
    if request.method == 'POST':

        title = request.POST.get('title', '')
        category = request.POST.get('category', '')
        content = request.POST.get('content', '')
        ids = request.POST.getlist('users[]')

        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        post = Post()
        post.member_id = admin.pk
        post.title = title
        post.category = category
        post.content = content
        post.picture_url = ''
        post.comments = '0'
        post.likes = '0'
        post.posted_time = str(int(round(time.time() * 1000)))
        post.save()

        fs = FileSystemStorage()
        i = 0
        for f in request.FILES.getlist('pictures'):
            i = i + 1
            filename = fs.save(f.name, f)
            uploaded_url = fs.url(filename)
            if i == 1:
                post.picture_url = settings.URL + uploaded_url
                post.save()
            postPicture = PostPicture()
            postPicture.post_id = post.pk
            postPicture.picture_url = settings.URL + uploaded_url
            postPicture.save()


        for member_id in ids:
            members = Member.objects.filter(id=int(member_id))
            if members.count() > 0:
                member = members[0]

                title = 'VaCay Community'
                subject = 'You\'ve received a post from VaCay Community'
                msg = 'Dear ' + member.name + ', You\'ve received a post from VaCay Community.<br><br>'
                msg = msg + '<a href=\'' + settings.URL + '/users/to_post?post_id=' + str(post.pk) + '\' target=\'_blank\'>View the post</a>'

                from_email = admin.email
                to_emails = []
                to_emails.append(member.email)
                send_mail_message(from_email, to_emails, title, subject, msg)

                msg = member.name + ', You\'ve received a message from VaCay Community.\n\n'
                msg = msg + 'Click on this link to view the post: ' + settings.URL + '/users/to_post?post_id=' + str(post.pk)

                notification = Notification()
                notification.member_id = member.pk
                notification.sender_id = admin.pk
                notification.message = msg
                notification.notified_time = str(int(round(time.time() * 1000)))
                notification.save()

                rcv = Received()
                rcv.member_id = member.pk
                rcv.sender_id = admin.pk
                rcv.noti_id = notification.pk
                rcv.save()

                snt = Sent()
                snt.member_id = member.pk
                snt.sender_id = admin.pk
                snt.noti_id = notification.pk
                snt.save()

                ##########################################################################################################################################################################

                db = firebase.database()
                data = {
                    "msg": msg,
                    "date":str(int(round(time.time() * 1000))),
                    "sender_id": str(admin.pk),
                    "sender_name": admin.name,
                    "sender_email": admin.email,
                    "sender_photo": admin.photo_url,
                    "role": "admin",
                    "type": "post",
                    "id": str(post.pk),
                    "mes_id": str(notification.pk)
                }

                db.child("notify").child(str(member.pk)).push(data)
                db.child("notify2").child(str(member.pk)).push(data)

                sendFCMPushNotification(member.pk, admin.pk, msg)

                #################################################################################################################################################################################

                if member.playerID != '':
                    playerIDList = []
                    playerIDList.append(member.playerID)
                    url = '/users/notifications?noti_id=' + str(notification.pk)
                    send_push(playerIDList, msg, url)

        return redirect('/users/posts/')



def add_post_comment(request):

    import datetime

    post_id = request.GET['post_id']

    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    members = Member.objects.filter(admin_id=admin.pk)

    posts = Post.objects.filter(id=post_id)
    if posts.count() == 0: return redirect('/posts/')

    post = posts[0]
    post.posted_time = datetime.datetime.fromtimestamp(float(int(post.posted_time)/1000)).strftime("%b %d, %Y %H:%M")

    comments = Comment.objects.filter(post_id=post.pk)
    post.comments = str(comments.count())
    likes = PostLike.objects.filter(post_id=post.pk)
    post.likes = str(likes.count())

    # return HttpResponse(post.member_id + '///' + str(admin.pk))

    if int(post.member_id) != admin.pk:

        pls = PostLike.objects.filter(post_id=post.pk, member_id=admin.pk)
        if pls.count() > 0: post.liked = 'yes'
        else: post.liked = 'no'

        ppictures = PostPicture.objects.filter(post_id=post.pk)

        comments = Comment.objects.filter(post_id=post_id, member_id=admin.pk)
        if comments.count() == 0:
            comments = Comment.objects.filter(post_id=post_id).order_by('-id')
            commentList = []
            for comment in comments:
                comment.commented_time = datetime.datetime.fromtimestamp(float(int(comment.commented_time)/1000)).strftime("%b %d, %Y %H:%M")
                members = Member.objects.filter(id=comment.member_id)
                if members.count() > 0:
                    member = members[0]
                    data = {
                        'comment':comment,
                        'member':member
                    }
                    commentList.append(data)
            members = Member.objects.filter(id=post.member_id)
            if members.count() == 0: return redirect('/posts/')
            member = members[0]
            data = {
                'post': post,
                'member': member,
                'pictures':ppictures
            }
            return render(request, 'motherwise/comment.html', {'post':data, 'me':admin, 'comments':commentList})

        else:
            myComment = comments[0]
            comments = Comment.objects.filter(post_id=post_id).order_by('-id')
            commentList = []
            for comment in comments:
                comment.commented_time = datetime.datetime.fromtimestamp(float(int(comment.commented_time)/1000)).strftime("%b %d, %Y %H:%M")
                members = Member.objects.filter(id=comment.member_id)
                if members.count() > 0:
                    member = members[0]
                    data = {
                        'comment':comment,
                        'member':member
                    }
                    commentList.append(data)
            members = Member.objects.filter(id=post.member_id)
            if members.count() == 0: return redirect('/posts/')
            member = members[0]
            data = {
                'post': post,
                'member': member,
                'pictures':ppictures
            }
            return render(request, 'motherwise/comment.html', {'post':data, 'me':admin, 'comments':commentList, 'comment':myComment})

    else:
        ppictures = PostPicture.objects.filter(post_id=post.pk)
        comments = Comment.objects.filter(post_id=post_id).order_by('-id')
        commentList = []
        for comment in comments:
            comment.commented_time = datetime.datetime.fromtimestamp(float(int(comment.commented_time)/1000)).strftime("%b %d, %Y %H:%M")
            members = Member.objects.filter(id=comment.member_id)
            if members.count() > 0:
                member = members[0]
                data = {
                    'comment':comment,
                    'member':member
                }
                commentList.append(data)
        data = {
            'post': post,
            'pictures':ppictures,
            'comments':commentList
        }
        return render(request, 'motherwise/edit_post.html', {'post':data, 'me':admin})




def delete_post_picture(request):
    picture_id = request.GET['picture_id']
    post_id = request.GET['post_id']
    posts = Post.objects.filter(id=post_id)
    if posts.count() == 0:
        return redirect('/posts/')
    post = posts[0]
    pics = PostPicture.objects.filter(id=picture_id, post_id=post_id)
    fs = FileSystemStorage()
    if pics.count() > 0:
        pic = pics[0]
        if pic.picture_url != '':
            fs.delete(pic.picture_url.replace(settings.URL + '/media/', ''))
            if pic.picture_url == post.picture_url:
                post.picture_url = ''
                pics = PostPicture.objects.filter(post_id=post_id)
                if pics.count() > 0:
                    post.picture_url = pics[0].picture_url
                post.save()
        pic.delete()
    return redirect('/add_post_comment?post_id=' + post_id)




def like_post(request):
    post_id = request.GET['post_id']

    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    posts = Post.objects.filter(id=post_id)
    if posts.count() == 0: return redirect('/posts/')

    post = posts[0]
    pls = PostLike.objects.filter(post_id=post.pk, member_id=admin.pk)
    if pls.count() > 0:
        pls[0].delete()
        post.likes = str(int(post.likes) - 1)
        post.save()
    else:
        pl = PostLike()
        pl.post_id = post.pk
        pl.member_id = admin.pk
        pl.liked_time = str(int(round(time.time() * 1000)))
        pl.save()

        post.likes = str(int(post.likes) + 1)
        post.save()

    # return redirect('/add_post_comment?post_id=' + str(post.pk))
    return HttpResponse(post.likes)



@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def submit_comment(request):
    if request.method == 'POST':

        post_id = request.POST.get('post_id', '')
        content = request.POST.get('content', '')

        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        fs = FileSystemStorage()

        comments = Comment.objects.filter(post_id=post_id, member_id=admin.pk)
        if comments.count() == 0:
            comment = Comment()
            comment.post_id = post_id
            comment.member_id = admin.pk
            comment.comment_text = content
            comment.commented_time = str(int(round(time.time() * 1000)))

            try:
                image = request.FILES['image']
                filename = fs.save(image.name, image)
                uploaded_url = fs.url(filename)
                comment.image_url = settings.URL + uploaded_url
            except MultiValueDictKeyError:
                print('no video updated')

            comment.save()

            post = Post.objects.get(id=post_id)
            post.comments = str(int(post.comments) + 1)
            post.save()

        else:
            comment = comments[0]
            comment.comment_text = content

            try:
                image = request.FILES['image']
                filename = fs.save(image.name, image)
                uploaded_url = fs.url(filename)
                if comment.image_url != '':
                    fs.delete(comment.image_url.replace(settings.URL + '/media/', ''))
                comment.image_url = settings.URL + uploaded_url
            except MultiValueDictKeyError:
                print('no video updated')

            comment.save()

        return redirect('/add_post_comment?post_id=' + post_id)


def delete_post(request):
    post_id = request.GET['post_id']

    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    fs = FileSystemStorage()

    posts = Post.objects.filter(id=post_id)
    if posts.count() == 0: return redirect('/posts/')

    post = posts[0]
    pls = PostLike.objects.filter(post_id=post.pk)
    for pl in pls:
        pl.delete()
    pps = PostPicture.objects.filter(post_id=post.pk)
    for pp in pps:
        if pp.picture_url != '':
            fs.delete(pp.picture_url.replace(settings.URL + '/media/', ''))
        pp.delete()
    pcs = Comment.objects.filter(post_id=post.pk)
    for pc in pcs:
        if pc.image_url != '':
            fs.delete(pc.image_url.replace(settings.URL + '/media/', ''))
        pc.delete()

    post.delete()

    return redirect('/posts/')



def delete_comment(request):
    comment_id = request.GET['comment_id']

    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    fs = FileSystemStorage()

    pcs = Comment.objects.filter(id=comment_id)
    if pcs.count() > 0:
        pc = pcs[0]
        post_id = pc.post_id
        if pc.image_url != '':
            fs.delete(pc.image_url.replace(settings.URL + '/media/', ''))
        pc.delete()

        post = Post.objects.get(id=post_id)
        post.comments = str(int(post.comments) - 1)
        post.save()

        return redirect('/add_post_comment?post_id=' + post_id)
    else:
        return redirect('/posts/')



@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def edit_post(request):
    if request.method == 'POST':

        post_id = request.POST.get('post_id', '1')
        title = request.POST.get('title', '')
        category = request.POST.get('category', '')
        content = request.POST.get('content', '')

        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        posts = Post.objects.filter(id=post_id)
        if posts.count() == 0:
            return redirect('/posts/')
        post = posts[0]
        post.title = title
        post.category = category
        post.content = content
        post.save()

        fs = FileSystemStorage()
        i = 0
        for f in request.FILES.getlist('pictures'):
            i = i + 1
            filename = fs.save(f.name, f)
            uploaded_url = fs.url(filename)
            if i == 1:
                post.picture_url = settings.URL + uploaded_url
                post.save()
            postPicture = PostPicture()
            postPicture.post_id = post.pk
            postPicture.picture_url = settings.URL + uploaded_url
            postPicture.save()

        return redirect('/add_post_comment?post_id=' + post_id)



@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def search_post(request):

    import datetime

    if request.method == 'POST':
        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        users = Member.objects.filter(admin_id=admin.pk).order_by('-id')
        userList = []
        for user in users:
            if user.registered_time != '':
                user.username = '@' + user.email[0:user.email.find('@')]
                userList.append(user)

        search_id = request.POST.get('q', None)

        posts = Post.objects.all().order_by('-id')
        postList = get_filtered_posts_data(admin, posts, search_id)

        list1 = []
        list2 = []
        list3 = []
        list4 = []

        i = 0
        for post in postList:
            post.posted_time = datetime.datetime.fromtimestamp(float(int(post.posted_time)/1000)).strftime("%b %d, %Y %H:%M")
            i = i + 1
            pls = PostLike.objects.filter(post_id=post.pk, member_id=admin.pk)
            if pls.count() > 0:
                post.liked = 'yes'
            else: post.liked = 'no'

            comments = Comment.objects.filter(post_id=post.pk)
            post.comments = str(comments.count())
            likes = PostLike.objects.filter(post_id=post.pk)
            post.likes = str(likes.count())

            member = Member.objects.get(id=post.member_id)

            data = {
                'member':member,
                'post': post
            }

            if i % 4 == 1: list1.append(data)
            elif i % 4 == 2: list2.append(data)
            elif i % 4 == 3: list3.append(data)
            elif i % 4 == 0: list4.append(data)

        return render(request, 'motherwise/post.html', {'me':admin, 'list1':list1, 'list2':list2, 'list3':list3, 'list4':list4, 'search':'Searched', 'users':userList})


# import datetime
from datetime import datetime

def get_filtered_posts_data(me, posts, keyword):
    postList = []
    for post in posts:
        members = Member.objects.filter(id=post.member_id)
        if members.count() > 0:
            member = members[0]
            if int(member.admin_id) == me.pk or int(post.member_id) == me.pk:
                if keyword.lower() in post.title.lower():
                    postList.append(post)
                elif keyword.lower() in post.category.lower():
                    postList.append(post)
                elif keyword.lower() in post.content.lower():
                    postList.append(post)
                elif keyword.lower() in post.comments.lower():
                    postList.append(post)
                elif keyword.lower() in post.likes.lower():
                    postList.append(post)
                elif keyword.lower() in member.name.lower():
                    postList.append(post)
                elif keyword.lower() in member.email.lower():
                    postList.append(post)
                elif keyword.lower() in member.phone_number.lower():
                    postList.append(post)
                elif keyword.lower() in member.cohort.lower():
                    postList.append(post)
                elif keyword.lower() in member.address.lower():
                    postList.append(post)
                else:
                    if keyword.isdigit():
                        keyDateObj = datetime.fromtimestamp(int(keyword)/1000)
                        postDateObj = datetime.fromtimestamp(int(post.posted_time)/1000)
                        if keyDateObj.year == postDateObj.year and keyDateObj.month == postDateObj.month and keyDateObj.day == postDateObj.day:
                            postList.append(post)

    return postList



def filter(request):

    import datetime

    option = request.GET['option']
    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    users = Member.objects.filter(admin_id=admin.pk).order_by('-id')
    userList = []
    for user in users:
        if user.registered_time != '':
            user.username = '@' + user.email[0:user.email.find('@')]
            userList.append(user)

    list1 = []
    list2 = []
    list3 = []
    list4 = []

    search = 'Searched'

    allPosts = Post.objects.all().order_by('-id')
    i = 0
    for post in allPosts:
        i = i + 1
        pls = PostLike.objects.filter(post_id=post.pk, member_id=admin.pk)
        if pls.count() > 0:
            post.liked = 'yes'
        else: post.liked = 'no'

        comments = Comment.objects.filter(post_id=post.pk)
        post.comments = str(comments.count())
        likes = PostLike.objects.filter(post_id=post.pk)
        post.likes = str(likes.count())

        members = Member.objects.filter(id=post.member_id)
        if members.count() > 0:
            memb = members[0]
            if int(memb.admin_id) == admin.pk or memb.pk == admin.pk:
                if option == 'last3':
                    if int(round(time.time() * 1000)) - int(post.posted_time) < 3 * 86400 * 1000:
                        post.posted_time = datetime.datetime.fromtimestamp(float(int(post.posted_time)/1000)).strftime("%b %d, %Y %H:%M")
                        data = {
                            'member':memb,
                            'post': post
                        }
                        if i % 4 == 1: list1.append(data)
                        elif i % 4 == 2: list2.append(data)
                        elif i % 4 == 3: list3.append(data)
                        elif i % 4 == 0: list4.append(data)
                        search = 'Last 3 Days'
                elif option == 'last7':
                    if int(round(time.time() * 1000)) - int(post.posted_time) < 7 * 86400 * 1000:
                        post.posted_time = datetime.datetime.fromtimestamp(float(int(post.posted_time)/1000)).strftime("%b %d, %Y %H:%M")
                        data = {
                            'member':memb,
                            'post': post
                        }
                        if i % 4 == 1: list1.append(data)
                        elif i % 4 == 2: list2.append(data)
                        elif i % 4 == 3: list3.append(data)
                        elif i % 4 == 0: list4.append(data)
                        search = 'Last 7 Days'
                elif option == 'last30':
                    if int(round(time.time() * 1000)) - int(post.posted_time) < 30 * 86400 * 1000:
                        post.posted_time = datetime.datetime.fromtimestamp(float(int(post.posted_time)/1000)).strftime("%b %d, %Y %H:%M")
                        data = {
                            'member':memb,
                            'post': post
                        }
                        if i % 4 == 1: list1.append(data)
                        elif i % 4 == 2: list2.append(data)
                        elif i % 4 == 3: list3.append(data)
                        elif i % 4 == 0: list4.append(data)
                        search = 'Last 30 Days'

    return render(request, 'motherwise/post.html', {'me':admin, 'list1':list1, 'list2':list2, 'list3':list3, 'list4':list4, 'search':search, 'users':userList})



def send_push(playerIDs, message, url):

    client = PybossaOneSignal(api_key=settings.OS_API_KEY, app_id=settings.OS_APP_ID)
    contents = {"en": message}
    headings = {"en": "VaCay Network"}
    launch_url = settings.URL + url
    chrome_web_image = settings.URL + '/static/images/notimage.jpg'
    chrome_web_icon = settings.URL + '/static/images/noticon.png'
    included_segments = []
    include_player_ids = playerIDs
    web_buttons=[{"id": "read-more-button",
                               "text": "Read more",
                               "icon": "http://i.imgur.com/MIxJp1L.png",
                               "url": launch_url}]
    try:
        client.push_msg(contents=contents, headings=headings, include_player_ids=include_player_ids, launch_url=launch_url, chrome_web_image=chrome_web_image, chrome_web_icon=chrome_web_icon, included_segments=included_segments, web_buttons=web_buttons)
    except:
        print('Error')



def notifications(request):

    import datetime

    noti_id = '0'

    try:
        noti_id = request.GET['noti_id']
    except MultiValueDictKeyError:
        print('No key')

    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    list1 = []
    list2 = []
    list3 = []
    list4 = []

    notis = Received.objects.filter(member_id=admin.pk).order_by('-id')

    i = 0
    for noti in notis:
        i = i + 1
        members = Member.objects.filter(id=noti.sender_id)
        if members.count() > 0:
            sender = members[0]
            nfs = Notification.objects.filter(id=noti.noti_id)
            if nfs.count() > 0:
                notification = nfs[0]
                notification.notified_time = datetime.datetime.fromtimestamp(float(int(notification.notified_time)/1000)).strftime("%b %d, %Y %H:%M")
                data = {
                    'sender':sender,
                    'noti': notification
                }

                if i % 4 == 1: list1.append(data)
                elif i % 4 == 2: list2.append(data)
                elif i % 4 == 3: list3.append(data)
                elif i % 4 == 0: list4.append(data)

    return render(request, 'motherwise/notifications.html', {'notid':noti_id, 'list1':list1, 'list2':list2, 'list3':list3, 'list4':list4, 'opt':'received'})




def sentnotis(request):

    import datetime

    noti_id = '0'

    try:
        noti_id = request.GET['noti_id']
    except MultiValueDictKeyError:
        print('No key')

    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    list1 = []
    list2 = []
    list3 = []
    list4 = []

    notis = Sent.objects.filter(sender_id=admin.pk).order_by('-id')

    # return HttpResponse(str(admin.pk) + '///' + str(notis.count()))

    i = 0
    for noti in notis:

        i = i + 1
        members = Member.objects.filter(id=noti.member_id)
        if members.count() > 0:
            receiver = members[0]
            nfs = Notification.objects.filter(id=noti.noti_id)
            if nfs.count() > 0:
                notification = nfs[0]
                notification.notified_time = datetime.datetime.fromtimestamp(float(int(notification.notified_time)/1000)).strftime("%b %d, %Y %H:%M")
                data = {
                    'receiver':receiver,
                    'noti': notification
                }

                if i % 4 == 1: list1.append(data)
                elif i % 4 == 2: list2.append(data)
                elif i % 4 == 3: list3.append(data)
                elif i % 4 == 0: list4.append(data)

    return render(request, 'motherwise/sent_notis.html', {'notid':noti_id, 'list1':list1, 'list2':list2, 'list3':list3, 'list4':list4, 'opt':'sent'})



def delete_noti(request):
    noti_id = request.GET['noti_id']
    opt = request.GET['opt']

    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    if opt == 'received':
        notis = Received.objects.filter(noti_id=noti_id)
        if notis.count() > 0:
            noti = notis[0]
            noti.delete()
        return redirect('/notifications/')
    elif opt == 'sent':
        notis = Sent.objects.filter(noti_id=noti_id)
        if notis.count() > 0:
            noti = notis[0]
            noti.delete()
        return redirect('/sentnotis/')



@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def processnewmessage(request):
    if request.method == 'POST':
        noti_id = request.POST.get('noti_id', '1')
        notis = Notification.objects.filter(id=noti_id)
        if notis.count() > 0:
            noti = notis[0]
            noti.status = 'read'
            noti.save()
        return HttpResponse('The message read')


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def notisearch(request):

    import datetime

    if request.method == 'POST':
        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        search_id = request.POST.get('q', None)
        opt = request.POST.get('opt', '')

        notis = []
        if opt == 'received':
            notis = Received.objects.filter(member_id=admin.pk).order_by('-id')
        elif opt == 'sent':
            notis = Sent.objects.filter(sender_id=admin.pk).order_by('-id')
        notiList = []
        for noti in notis:
            nfs = Notification.objects.filter(id=noti.noti_id)
            if nfs.count() > 0:
                notification = nfs[0]
                notiList.append(notification)
        notiList = get_filtered_notis_data(notiList, search_id, opt)

        list1 = []
        list2 = []
        list3 = []
        list4 = []

        i = 0
        for noti in notiList:
            noti.notified_time = datetime.datetime.fromtimestamp(float(int(noti.notified_time)/1000)).strftime("%b %d, %Y %H:%M")
            i = i + 1
            members = []
            if opt == 'received':
                members = Member.objects.filter(id=noti.sender_id)
            elif opt == 'sent':
                members = Member.objects.filter(id=noti.member_id)
            if members.count() > 0:
                member = members[0]
                data = {}
                if opt == 'received':
                    data = {
                        'sender':member,
                        'noti': noti
                    }
                elif opt == 'sent':
                    data = {
                        'receiver':member,
                        'noti': noti
                    }

                if i % 4 == 1: list1.append(data)
                elif i % 4 == 2: list2.append(data)
                elif i % 4 == 3: list3.append(data)
                elif i % 4 == 0: list4.append(data)

        if opt == 'sent':
            return render(request, 'motherwise/sent_notis.html', {'list1':list1, 'list2':list2, 'list3':list3, 'list4':list4, 'search':'Searched', 'opt':'sent'})

        return render(request, 'motherwise/notifications.html', {'list1':list1, 'list2':list2, 'list3':list3, 'list4':list4, 'search':'Searched', 'opt':'received'})


# import datetime
# from datetime import datetime

def get_filtered_notis_data(notis, keyword, option):
    notiList = []
    for noti in notis:
        members = []
        if option == 'received':
            members = Member.objects.filter(id=noti.sender_id)
        elif option == 'sent':
            members = Member.objects.filter(id=noti.member_id)
        if members.count() > 0:
            member = members[0]
            if keyword.lower() in noti.message.lower():
                notiList.append(noti)
            elif keyword.lower() in member.name.lower():
                notiList.append(noti)
            elif keyword.lower() in member.email.lower():
                notiList.append(noti)
            elif keyword.lower() in member.phone_number.lower():
                notiList.append(noti)
            elif keyword.lower() in member.cohort.lower():
                notiList.append(noti)
            elif keyword.lower() in member.address.lower():
                notiList.append(noti)
            else:
                if keyword.isdigit():
                    keyDateObj = datetime.fromtimestamp(int(keyword)/1000)
                    notiDateObj = datetime.fromtimestamp(int(noti.notified_time)/1000)
                    if keyDateObj.year == notiDateObj.year and keyDateObj.month == notiDateObj.month and keyDateObj.day == notiDateObj.day:
                        notiList.append(noti)

    return notiList



def fffff(request):

    import datetime

    option = request.GET['noption']
    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    list1 = []
    list2 = []
    list3 = []
    list4 = []

    search = 'Searched'
    opt = request.GET['opt']

    notis = []
    if opt == 'received':
        notis = Received.objects.filter(member_id=admin.pk).order_by('-id')
    elif opt == 'sent':
        notis = Sent.objects.filter(sender_id=admin.pk).order_by('-id')

    notiList = []
    for noti in notis:
        nfs = Notification.objects.filter(id=noti.noti_id)
        if nfs.count() > 0:
            notification = nfs[0]
            notiList.append(notification)

    i = 0
    for noti in notiList:
        i = i + 1
        if opt == 'received':
            members = Member.objects.filter(id=noti.sender_id)
            if members.count() > 0:
                sender = members[0]
                if option == 'last3':
                    if int(round(time.time() * 1000)) - int(noti.notified_time) < 3 * 86400 * 1000:
                        noti.notified_time = datetime.datetime.fromtimestamp(float(int(noti.notified_time)/1000)).strftime("%b %d, %Y %H:%M")
                        data = {
                            'sender':sender,
                            'noti': noti
                        }
                        if i % 4 == 1: list1.append(data)
                        elif i % 4 == 2: list2.append(data)
                        elif i % 4 == 3: list3.append(data)
                        elif i % 4 == 0: list4.append(data)
                        search = 'Last 3 Days'
                elif option == 'last7':
                    if int(round(time.time() * 1000)) - int(noti.notified_time) < 7 * 86400 * 1000:
                        noti.notified_time = datetime.datetime.fromtimestamp(float(int(noti.notified_time)/1000)).strftime("%b %d, %Y %H:%M")
                        data = {
                            'sender':sender,
                            'noti': noti
                        }
                        if i % 4 == 1: list1.append(data)
                        elif i % 4 == 2: list2.append(data)
                        elif i % 4 == 3: list3.append(data)
                        elif i % 4 == 0: list4.append(data)
                        search = 'Last 7 Days'
                elif option == 'last30':
                    if int(round(time.time() * 1000)) - int(noti.notified_time) < 30 * 86400 * 1000:
                        noti.notified_time = datetime.datetime.fromtimestamp(float(int(noti.notified_time)/1000)).strftime("%b %d, %Y %H:%M")
                        data = {
                            'sender':sender,
                            'noti': noti
                        }
                        if i % 4 == 1: list1.append(data)
                        elif i % 4 == 2: list2.append(data)
                        elif i % 4 == 3: list3.append(data)
                        elif i % 4 == 0: list4.append(data)
                        search = 'Last 30 Days'
        elif opt == 'sent':
            members = Member.objects.filter(id=noti.member_id)
            if members.count() > 0:
                receiver = members[0]
                if option == 'last3':
                    if int(round(time.time() * 1000)) - int(noti.notified_time) < 3 * 86400 * 1000:
                        noti.notified_time = datetime.datetime.fromtimestamp(float(int(noti.notified_time)/1000)).strftime("%b %d, %Y %H:%M")
                        data = {
                            'receiver':receiver,
                            'noti': noti
                        }
                        if i % 4 == 1: list1.append(data)
                        elif i % 4 == 2: list2.append(data)
                        elif i % 4 == 3: list3.append(data)
                        elif i % 4 == 0: list4.append(data)
                        search = 'Last 3 Days'
                elif option == 'last7':
                    if int(round(time.time() * 1000)) - int(noti.notified_time) < 7 * 86400 * 1000:
                        noti.notified_time = datetime.datetime.fromtimestamp(float(int(noti.notified_time)/1000)).strftime("%b %d, %Y %H:%M")
                        data = {
                            'receiver':receiver,
                            'noti': noti
                        }
                        if i % 4 == 1: list1.append(data)
                        elif i % 4 == 2: list2.append(data)
                        elif i % 4 == 3: list3.append(data)
                        elif i % 4 == 0: list4.append(data)
                        search = 'Last 7 Days'
                elif option == 'last30':
                    if int(round(time.time() * 1000)) - int(noti.notified_time) < 30 * 86400 * 1000:
                        noti.notified_time = datetime.datetime.fromtimestamp(float(int(noti.notified_time)/1000)).strftime("%b %d, %Y %H:%M")
                        data = {
                            'receiver':receiver,
                            'noti': noti
                        }
                        if i % 4 == 1: list1.append(data)
                        elif i % 4 == 2: list2.append(data)
                        elif i % 4 == 3: list3.append(data)
                        elif i % 4 == 0: list4.append(data)
                        search = 'Last 30 Days'

    if opt == 'sent':
        return render(request, 'motherwise/sent_notis.html', {'list1':list1, 'list2':list2, 'list3':list3, 'list4':list4, 'search':'Searched', 'opt':'sent'})

    return render(request, 'motherwise/notifications.html', {'list1':list1, 'list2':list2, 'list3':list3, 'list4':list4, 'search':search, 'opt':'received'})




def videotest(request):
    return render(request, 'motherwise/video_agora.html')


def open_conference(request):

    import datetime

    group_id = request.GET['group_id']
    cohort = request.GET['cohort']

    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    memberList = []
    if group_id != '' and int(group_id) > 0:
        group = None
        groups = Group.objects.filter(member_id=admin.pk, id=group_id).order_by('-id')
        if groups.count() > 0:
            group = groups[0]
            gMembers = GroupMember.objects.filter(group_id=group.pk)
            for gMember in gMembers:
                members = Member.objects.filter(id=gMember.member_id)
                if members.count() > 0:
                    memberList.append(members[0])

            request.session['group_id'] = group_id
            request.session['cohort'] = ''

            memberIdList = []
            for memb in memberList:
                memberIdList.append(memb.pk)
            request.session['selected_member_list'] = memberIdList

            confs = Conference.objects.filter(member_id=admin.pk, group_id=group_id).order_by('-id')
            for conf in confs:
                conf.created_time = datetime.datetime.fromtimestamp(float(int(conf.created_time)/1000)).strftime("%b %d, %Y %H:%M")
                if conf.event_time != '': conf.event_time = datetime.datetime.fromtimestamp(float(int(conf.event_time)/1000)).strftime("%b %d, %Y %H:%M")
            if confs.count() == 0:
                return render(request, 'motherwise/conference_create.html', {'conf_opt':'new_conference', 'confs':confs, 'group':group})
            try:
                option = request.GET['option']
                if option == 'new_conference':
                    return render(request, 'motherwise/conference_create.html', {'conf_opt':'new_conference', 'confs':confs, 'group':group})
            except KeyError:
                print('no key')
            # return render(request, 'motherwise/conference_create.html', {'conf_opt':'new_conference', 'confs':confs, 'group':group})

            last_conf = confs[0]

            try:
                conf_id = request.GET['conf_id']
                cfs = Conference.objects.filter(id=conf_id)
                if cfs.count() > 0:
                    last_conf = cfs[0]
                    last_conf.created_time = datetime.datetime.fromtimestamp(float(int(last_conf.created_time)/1000)).strftime("%b %d, %Y %H:%M")
                    if last_conf.event_time != '': last_conf.event_time = datetime.datetime.fromtimestamp(float(int(last_conf.event_time)/1000)).strftime("%b %d, %Y %H:%M")
            except KeyError:
                print('no key')

            if last_conf.type == 'live':
                return render(request, 'motherwise/conf_live.html', {'me':admin, 'members':memberList, 'group':group, 'confs':confs, 'last_conf':last_conf})
            elif last_conf.type == 'file':
                return render(request, 'motherwise/conf_video.html', {'me':admin, 'members':memberList, 'group':group, 'confs':confs, 'last_conf':last_conf})
            elif last_conf.type == 'youtube':
                return render(request, 'motherwise/conf_youtube.html', {'me':admin, 'members':memberList, 'group':group, 'confs':confs, 'last_conf':last_conf})
            else:
                return redirect('/home')

    elif cohort != '':
        members = Member.objects.filter(admin_id=admin.pk, cohort=cohort)
        if members.count() > 0:
            for member in members:
                memberList.append(member)

            request.session['group_id'] = ''
            request.session['cohort'] = cohort

            memberIdList = []
            for memb in memberList:
                memberIdList.append(memb.pk)
            request.session['selected_member_list'] = memberIdList

            confs = Conference.objects.filter(member_id=admin.pk, cohort=cohort).order_by('-id')
            for conf in confs:
                conf.created_time = datetime.datetime.fromtimestamp(float(int(conf.created_time)/1000)).strftime("%b %d, %Y %H:%M")
                if conf.event_time != '': conf.event_time = datetime.datetime.fromtimestamp(float(int(conf.event_time)/1000)).strftime("%b %d, %Y %H:%M")
            if confs.count() == 0:
                return render(request, 'motherwise/conference_create.html', {'conf_opt':'new_conference', 'confs':confs, 'cohort':cohort})
            try:
                option = request.GET['option']
                if option == 'new_conference':
                    return render(request, 'motherwise/conference_create.html', {'conf_opt':'new_conference', 'confs':confs, 'cohort':cohort})
            except KeyError:
                print('no session')
            # return render(request, 'motherwise/conference_create.html', {'conf_opt':'new_conference', 'confs':confs, 'group':group})

            last_conf = confs[0]

            try:
                conf_id = request.GET['conf_id']
                cfs = Conference.objects.filter(id=conf_id)
                if cfs.count() > 0:
                    last_conf = cfs[0]
                    last_conf.created_time = datetime.datetime.fromtimestamp(float(int(last_conf.created_time)/1000)).strftime("%b %d, %Y %H:%M")
                    if last_conf.event_time != '': last_conf.event_time = datetime.datetime.fromtimestamp(float(int(last_conf.event_time)/1000)).strftime("%b %d, %Y %H:%M")
            except KeyError:
                print('no key')

            if last_conf.type == 'live':
                return render(request, 'motherwise/conf_live.html', {'me':admin, 'members':memberList, 'cohort':cohort, 'confs':confs, 'last_conf':last_conf})
            elif last_conf.type == 'file':
                return render(request, 'motherwise/conf_video.html', {'me':admin, 'members':memberList, 'cohort':cohort, 'confs':confs, 'last_conf':last_conf})
            elif last_conf.type == 'youtube':
                return render(request, 'motherwise/conf_youtube.html', {'me':admin, 'members':memberList, 'cohort':cohort, 'confs':confs, 'last_conf':last_conf})
            else:
                return redirect('/home')


def genRandomConferenceCode():
    import strgen
    randomString = strgen.StringGenerator("[\w\d]{6}").render()
    return randomString


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def create_conference(request):

    import datetime

    if request.method == 'POST':

        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        group_id = request.POST.get('group_id', '1')
        cohort = request.POST.get('cohort', '')

        try:
            gtype = request.POST.get('gtype', '')
            if gtype == 'group':
                cohort = ''
            elif gtype == 'cohort':
                group_id = ''
        except KeyError:
            print('no key')
            return render(request, 'motherwise/admin.html')

        name = request.POST.get('name', '')
        type = request.POST.get('type', '')
        youtubeurl = request.POST.get('youtubeurl', '')
        event_time = request.POST.get('event_time', '')

        conf = Conference()
        conf.member_id = admin.pk
        conf.group_id = group_id
        conf.cohort = cohort
        conf.name = name
        if type == 'live':
            conf.code = genRandomConferenceCode()
        conf.type = type
        conf.event_time = event_time
        conf.created_time = str(int(round(time.time() * 1000)))
        conf.participants = '0'
        conf.likes = '0'

        dt = ''

        if type == 'youtube':
            conf.video_url = youtubeurl
            if event_time != '':
                dt = datetime.datetime.fromtimestamp(float(int(event_time)/1000)).strftime("%b %d, %Y %H:%M")
        elif type == 'file':
            try:
                videofile = request.FILES['video']
                fs = FileSystemStorage()
                filename = fs.save(videofile.name, videofile)
                video_url = fs.url(filename)
                conf.video_url = settings.URL + video_url
            except MultiValueDictKeyError:
                print("File Not Exist")

            if event_time != '':
                dt = datetime.datetime.fromtimestamp(float(int(event_time)/1000)).strftime("%b %d, %Y %H:%M")

        elif type == 'live':
            conf.video_url = ''
            dt = datetime.datetime.fromtimestamp(float(int(event_time)/1000)).strftime("%b %d, %Y %H:%M")

        conf.save()

        option = request.POST.get('option')

        if option == 'message':
            if group_id != '' and int(group_id) > 0:
                groups = Group.objects.filter(id=group_id)
                if groups.count() > 0:
                    group = groups[0]
                    gMembers = GroupMember.objects.filter(group_id=group.pk)
                    for gMember in gMembers:
                        members = Member.objects.filter(id=gMember.member_id)
                        if members.count() > 0:
                            member = members[0]

                            notification = Notification()
                            notification.member_id = member.pk
                            notification.sender_id = admin.pk
                            notification.notified_time = str(int(round(time.time() * 1000)))
                            notification.save()

                            rcv = Received()
                            rcv.member_id = member.pk
                            rcv.sender_id = admin.pk
                            rcv.noti_id = notification.pk
                            rcv.save()

                            snt = Sent()
                            snt.member_id = member.pk
                            snt.sender_id = admin.pk
                            snt.noti_id = notification.pk
                            snt.save()

                            title = 'Video Conference Invitation From VaCay Community'
                            subject = 'You\'ve received an invitation from VaCay Community'
                            msg = 'Dear ' + member.name + ',<br><br>You\'ve received an invitation to a video conference in a community ' + group.name + ' from VaCay Community.<br><br>Community name: ' + group.name + '<br>'
                            msg = msg + 'Conference topic: ' + name + '<br>'
                            msg = msg + 'Entry code: ' + conf.code + '<br>'
                            if type == 'live': type = 'Live'
                            elif type == 'file': type = 'General'
                            elif type == 'youtube': type = 'YouTube'
                            msg = msg + 'Type: ' + type + '<br>'
                            if dt != '':
                                msg = msg + 'Conference start date and time: ' + dt + '<br>'
                            msg = msg + 'So you can see this conference in the community and connect it to attend the event at that time.<br>'
                            msg = msg + 'Please join the community conference.'
                            msg = msg + '<br><a href=\'' + settings.URL + '/users/notifications?noti_id=' + str(notification.pk) + '\' target=\'_blank\'>Join website</a>' + '<br><br>VaCay Team'

                            from_email = admin.email
                            to_emails = []
                            to_emails.append(member.email)
                            send_mail_message(from_email, to_emails, title, subject, msg)

                            msg = 'Dear ' + member.name + ',\nYou\'ve received an invitation to a community ' + group.name + ' from VaCay Community.\nCommunity name: ' + group.name + '\n'
                            msg = msg + 'Conference topic: ' + name + '\n'
                            msg = msg + 'Entry code: ' + conf.code + '\n'
                            if type == 'live': type = 'Live'
                            elif type == 'file': type = 'General'
                            elif type == 'youtube': type = 'YouTube'
                            msg = msg + 'Type: ' + type + '\n'
                            if dt != '':
                                msg = msg + 'Conference start date and time: ' + dt + '\n'
                            msg = msg + 'So you can see this conference in the community and connect it to attend the event at that time.\n'
                            msg = msg + 'Please join the community conference.\nBest Regards\n\nVaCay Community'

                            notification.message = msg
                            notification.save()

                            ##########################################################################################################################################################################

                            db = firebase.database()
                            data = {
                                "msg": msg,
                                "date":str(int(round(time.time() * 1000))),
                                "sender_id": str(admin.pk),
                                "sender_name": admin.name,
                                "sender_email": admin.email,
                                "sender_photo": admin.photo_url,
                                "role": "admin",
                                "type": "conference",
                                "id": str(conf.pk),
                                "mes_id": str(notification.pk)
                            }

                            db.child("notify").child(str(member.pk)).push(data)
                            db.child("notify2").child(str(member.pk)).push(data)

                            sendFCMPushNotification(member.pk, admin.pk, msg)

                            #################################################################################################################################################################################

                            if member.playerID != '':
                                playerIDList = []
                                playerIDList.append(member.playerID)
                                url = '/users/notifications?noti_id=' + str(notification.pk)
                                send_push(playerIDList, msg, url)

                    conf.status = 'notified'
                    conf.save()

            elif cohort != '':
                members = Member.objects.filter(admin_id=admin.pk, cohort=cohort)
                if members.count() > 0:
                    for member in members:

                        notification = Notification()
                        notification.member_id = member.pk
                        notification.sender_id = admin.pk
                        notification.notified_time = str(int(round(time.time() * 1000)))
                        notification.save()

                        rcv = Received()
                        rcv.member_id = member.pk
                        rcv.sender_id = admin.pk
                        rcv.noti_id = notification.pk
                        rcv.save()

                        snt = Sent()
                        snt.member_id = member.pk
                        snt.sender_id = admin.pk
                        snt.noti_id = notification.pk
                        snt.save()

                        title = 'Video Conference Invitation From VaCay Community'
                        subject = 'You\'ve received an invitation from VaCay Community'
                        msg = 'Dear ' + member.name + ',<br><br>You\'ve received an invitation to a video conference in a group ' + cohort + ' from VaCay Community.<br><br>Group name: ' + cohort + '<br>'
                        msg = msg + 'Conference topic: ' + name + '<br>'
                        msg = msg + 'Entry code: ' + conf.code + '<br>'
                        if type == 'live': type = 'Live'
                        elif type == 'file': type = 'General'
                        elif type == 'youtube': type = 'YouTube'
                        msg = msg + 'Type: ' + type + '<br>'
                        if dt != '':
                            msg = msg + 'Conference start date and time: ' + dt + '<br>'
                        msg = msg + 'So you can see this conference in the group and connect it to attend the event at that time.<br>'
                        msg = msg + 'Please join the group conference.'
                        msg = msg + '<br><a href=\'' + settings.URL + '/users/notifications?noti_id=' + str(notification.pk) + '\' target=\'_blank\'>Join website</a>' + '<br><br>VaCay Team'

                        from_email = admin.email
                        to_emails = []
                        to_emails.append(member.email)
                        send_mail_message(from_email, to_emails, title, subject, msg)

                        msg = 'Dear ' + member.name + ',\nYou\'ve received an invitation to a group ' + cohort + ' from VaCay Community.\nGroup name: ' + cohort + '\n'
                        msg = msg + 'Conference topic: ' + name + '\n'
                        msg = msg + 'Entry code: ' + conf.code + '\n'
                        if type == 'live': type = 'Live'
                        elif type == 'file': type = 'General'
                        elif type == 'youtube': type = 'YouTube'
                        msg = msg + 'Type: ' + type + '\n'
                        if dt != '':
                            msg = msg + 'Conference start date and time: ' + dt + '\n'
                        msg = msg + 'So you can see this conference in the group and connect it to attend the event at that time.\n'
                        msg = msg + 'Please join the group conference.\nBest Regards\n\nVaCay Community'

                        notification.message = msg
                        notification.save()

                        ##########################################################################################################################################################################

                        db = firebase.database()
                        data = {
                            "msg": msg,
                            "date":str(int(round(time.time() * 1000))),
                            "sender_id": str(admin.pk),
                            "sender_name": admin.name,
                            "sender_email": admin.email,
                            "sender_photo": admin.photo_url,
                            "role": "admin",
                            "type": "conference",
                            "id": str(conf.pk),
                            "mes_id": str(notification.pk)
                        }

                        db.child("notify").child(str(member.pk)).push(data)
                        db.child("notify2").child(str(member.pk)).push(data)

                        sendFCMPushNotification(member.pk, admin.pk, msg)

                        #################################################################################################################################################################################

                        if member.playerID != '':
                            playerIDList = []
                            playerIDList.append(member.playerID)
                            url = '/users/notifications?noti_id=' + str(notification.pk)
                            send_push(playerIDList, msg, url)

                    conf.status = 'notified'
                    conf.save()

        else:
            print('no message')

        try:
            branch = request.POST.get('branch', '')
            if branch == 'conferences':
                return redirect('/to_conferences')
        except KeyError:
            print('no key')

        return redirect('/open_conference?conf_id=' + str(conf.pk) + '&group_id=' + group_id + '&cohort=' + cohort)



@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def delete_conference(request):
    if request.method == 'GET':
        conf_id = request.GET['conf_id']

        confs = Conference.objects.filter(id=conf_id)

        fs = FileSystemStorage()

        if confs.count() > 0:
            conf = confs[0]
            group_id = conf.group_id
            cohort = conf.cohort
            if conf.video_url != '' and 'http' in conf.video_url:
                fname = conf.video_url.replace(settings.URL + '/media/', '')
                fs.delete(fname)
            conf.delete()

            try:
                branch = request.GET['branch']
                if branch == 'conferences':
                    return redirect('/to_conferences')
            except KeyError:
                print('no key')

            return redirect('/open_conference?group_id=' + group_id + '&cohort=' + cohort)
        else:
            return redirect('/home')



@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def conference_notify(request):

    import datetime

    if request.method == 'POST':

        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        group_id = request.POST.get('group_id', '1')
        cohort = request.POST.get('cohort', '')
        conf_id = request.POST.get('conf_id', '1')

        confs = Conference.objects.filter(id=conf_id)
        if confs.count() == 0:
            return redirect('/home')

        conf = confs[0]

        name = conf.name
        type = conf.type
        event_time = conf.event_time

        dt = ''

        if event_time != '':

            dt = datetime.datetime.fromtimestamp(float(int(event_time)/1000)).strftime("%b %d, %Y %H:%M")

        if admin is not None:
            if group_id != '' and int(group_id) > 0:
                groups = Group.objects.filter(id=group_id)
                if groups.count() > 0:
                    group = groups[0]
                    gMembers = GroupMember.objects.filter(group_id=group.pk)
                    for gMember in gMembers:
                        members = Member.objects.filter(id=gMember.member_id)
                        if members.count() > 0:
                            member = members[0]

                            notification = Notification()
                            notification.member_id = member.pk
                            notification.sender_id = admin.pk
                            notification.notified_time = str(int(round(time.time() * 1000)))
                            notification.save()

                            rcv = Received()
                            rcv.member_id = member.pk
                            rcv.sender_id = admin.pk
                            rcv.noti_id = notification.pk
                            rcv.save()

                            snt = Sent()
                            snt.member_id = member.pk
                            snt.sender_id = admin.pk
                            snt.noti_id = notification.pk
                            snt.save()

                            title = 'Video Conference Invitation From VaCay Community'
                            subject = 'You\'ve received an invitation from VaCay Community'
                            msg = 'Dear ' + member.name + ',<br><br>You\'ve received an invitation to a video conference in a community ' + group.name + ' from VaCay Community.<br><br>Community name: ' + group.name + '<br>'
                            msg = msg + 'Conference topic: ' + conf.name + '<br>'
                            msg = msg + 'Entry code: ' + conf.code + '<br>'
                            type = conf.type
                            if type == 'live': type = 'Live'
                            elif type == 'file': type = 'General'
                            elif type == 'youtube': type = 'YouTube'
                            msg = msg + 'Type: ' + type + '<br>'
                            if dt != '':
                                msg = msg + 'Conference start date and time: ' + dt + '<br>'
                            msg = msg + 'So you can see this conference in the community and connect it to attend the event at that time.<br>'
                            msg = msg + 'Please join the community conference.'
                            msg = msg + '<br><a href=\'' + settings.URL + '/users/notifications?noti_id=' + str(notification.pk) + '\' target=\'_blank\'>Join website</a>' + '<br><br>VaCay Team'

                            from_email = admin.email
                            to_emails = []
                            to_emails.append(member.email)
                            send_mail_message(from_email, to_emails, title, subject, msg)

                            msg = 'Dear ' + member.name + ',\nYou\'ve received an invitation to a community ' + group.name + ' from VaCay Community.\nCommunity name: ' + group.name + '\n'
                            msg = msg + 'Conference topic: ' + conf.name + '\n'
                            msg = msg + 'Entry code: ' + conf.code + '\n'
                            type = conf.type
                            if type == 'live': type = 'Live'
                            elif type == 'file': type = 'General'
                            elif type == 'youtube': type = 'YouTube'
                            msg = msg + 'Type: ' + type + '\n'
                            if dt != '':
                                msg = msg + 'Conference start date and time: ' + dt + '\n'
                            msg = msg + 'So you can see this conference in the community and connect it to attend the event at that time.\n'
                            msg = msg + 'Please join the community conference.\nBest Regards\n\nVaCay Community'

                            notification.message = msg
                            notification.save()

                            ##########################################################################################################################################################################

                            db = firebase.database()
                            data = {
                                "msg": msg,
                                "date":str(int(round(time.time() * 1000))),
                                "sender_id": str(admin.pk),
                                "sender_name": admin.name,
                                "sender_email": admin.email,
                                "sender_photo": admin.photo_url,
                                "role": "admin",
                                "type": "conference",
                                "id": str(conf.pk),
                                "mes_id": str(notification.pk)
                            }

                            db.child("notify").child(str(member.pk)).push(data)
                            db.child("notify2").child(str(member.pk)).push(data)

                            sendFCMPushNotification(member.pk, admin.pk, msg)

                            #################################################################################################################################################################################

                            if member.playerID != '':
                                playerIDList = []
                                playerIDList.append(member.playerID)
                                url = '/users/notifications?noti_id=' + str(notification.pk)
                                send_push(playerIDList, msg, url)

                    conf.status = 'notified'
                    conf.save()

            elif cohort != '':
                members = Member.objects.filter(admin_id=admin.pk, cohort=cohort)
                if members.count() > 0:
                    for member in members:

                        notification = Notification()
                        notification.member_id = member.pk
                        notification.sender_id = admin.pk
                        notification.notified_time = str(int(round(time.time() * 1000)))
                        notification.save()

                        rcv = Received()
                        rcv.member_id = member.pk
                        rcv.sender_id = admin.pk
                        rcv.noti_id = notification.pk
                        rcv.save()

                        snt = Sent()
                        snt.member_id = member.pk
                        snt.sender_id = admin.pk
                        snt.noti_id = notification.pk
                        snt.save()

                        title = 'Video Conference Invitation From VaCay Community'
                        subject = 'You\'ve received an invitation from VaCay Community'
                        msg = 'Dear ' + member.name + ',<br><br>You\'ve received an invitation to a video conference in a group ' + cohort + ' from VaCay Community.<br><br>Group name: ' + cohort + '<br>'
                        msg = msg + 'Conference topic: ' + conf.name + '<br>'
                        msg = msg + 'Entry code: ' + conf.code + '<br>'
                        type = conf.type
                        if type == 'live': type = 'Live'
                        elif type == 'file': type = 'General'
                        elif type == 'youtube': type = 'YouTube'
                        msg = msg + 'Type: ' + type + '<br>'
                        if dt != '':
                            msg = msg + 'Conference start date and time: ' + dt + '<br>'
                        msg = msg + 'So you can see this conference in the group and connect it to attend the event at that time.<br>'
                        msg = msg + 'Please join the group conference.'
                        msg = msg + '<br><a href=\'' + settings.URL + '/users/notifications?noti_id=' + str(notification.pk) + '\' target=\'_blank\'>Join website</a>' + '<br><br>VaCay Team'

                        from_email = admin.email
                        to_emails = []
                        to_emails.append(member.email)
                        send_mail_message(from_email, to_emails, title, subject, msg)

                        msg = 'Dear ' + member.name + ',\nYou\'ve received an invitation to a group ' + cohort + ' from VaCay Community.\nGroup name: ' + cohort + '\n'
                        msg = msg + 'Conference topic: ' + conf.name + '\n'
                        msg = msg + 'Entry code: ' + conf.code + '\n'
                        type = conf.type
                        if type == 'live': type = 'Live'
                        elif type == 'file': type = 'General'
                        elif type == 'youtube': type = 'YouTube'
                        msg = msg + 'Type: ' + type + '\n'
                        if dt != '':
                            msg = msg + 'Conference start date and time: ' + dt + '\n'
                        msg = msg + 'So you can see this conference in the group and connect it to attend the event at that time.\n'
                        msg = msg + 'Please join the group conference.\nBest Regards\n\nVaCay Community'

                        notification.message = msg
                        notification.save()

                        ##########################################################################################################################################################################

                        db = firebase.database()
                        data = {
                            "msg": msg,
                            "date":str(int(round(time.time() * 1000))),
                            "sender_id": str(admin.pk),
                            "sender_name": admin.name,
                            "sender_email": admin.email,
                            "sender_photo": admin.photo_url,
                            "role": "admin",
                            "type": "conference",
                            "id": str(conf.pk),
                            "mes_id": str(notification.pk)
                        }

                        db.child("notify").child(str(member.pk)).push(data)
                        db.child("notify2").child(str(member.pk)).push(data)

                        sendFCMPushNotification(member.pk, admin.pk, msg)

                        #################################################################################################################################################################################

                        if member.playerID != '':
                            playerIDList = []
                            playerIDList.append(member.playerID)
                            url = '/users/notifications?noti_id=' + str(notification.pk)
                            send_push(playerIDList, msg, url)

                    conf.status = 'notified'
                    conf.save()

        else:
            return render(request, 'motherwise/admin.html')

        return redirect('/open_conference?conf_id=' + str(conf.pk) + '&group_id=' + group_id + '&cohort=' + cohort)




@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def video_selected_members(request):

    import datetime

    if request.method == 'POST':

        ids = request.POST.getlist('users[]')

        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        memberList = []
        memberList2 = []
        memberIdList = []
        for member_id in ids:
            members = Member.objects.filter(id=int(member_id))
            if members.count() > 0:
                member = members[0]
                memberList.append(member)
                memberIdList.append(member.pk)

        if len(memberList) > 0:
            request.session['selected_member_list'] = memberIdList
            groups = Group.objects.filter(member_id=admin.pk).order_by('-id')
            for group in groups:
                group.created_time = datetime.datetime.fromtimestamp(float(int(group.created_time)/1000)).strftime("%b %d, %Y %H:%M")
                group.last_connected_time = datetime.datetime.fromtimestamp(float(int(group.last_connected_time)/1000)).strftime("%b %d, %Y %H:%M")
            latestGroupMemberList = []
            latest_group = None
            if groups.count() > 0:
                latest_group = groups[0]
                gMembers = GroupMember.objects.filter(group_id=latest_group.pk)
                for gMember in gMembers:
                    members = Member.objects.filter(id=gMember.member_id)
                    if members.count() > 0:
                        latestGroupMemberList.append(members[0])
                for memb in memberList:
                    gms = GroupMember.objects.filter(group_id=latest_group.pk, member_id=memb.pk)
                    if gms.count() == 0: memberList2.append(memb)
            else:
                memberList2 = memberList
            gcs = GroupConnect.objects.filter(member_id=admin.pk).order_by('-id')
            recents = []
            for gc in gcs:
                gs = Group.objects.filter(id=gc.group_id)
                if gs.count() > 0:
                    group = gs[0]
                    group.created_time = datetime.datetime.fromtimestamp(float(int(group.created_time)/1000)).strftime("%b %d, %Y %H:%M")
                    group.last_connected_time = datetime.datetime.fromtimestamp(float(int(group.last_connected_time)/1000)).strftime("%b %d, %Y %H:%M")
                    recents.append(group)

            return render(request, 'motherwise/groups.html', {'members':memberList2, 'group':latest_group, 'groups': groups, 'group_members':latestGroupMemberList, 'recents':recents})
        else:
            return redirect('/home')

    else:
        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        memberIdList = []
        try:
            memberIdList = request.session['selected_member_list']
        except KeyError:
            print('No key')

        memberList = []
        for member_id in memberIdList:
            members = Member.objects.filter(id=member_id)
            if members.count() > 0:
                member = members[0]
                memberList.append(member)

        contacts = update_admin_contact(admin, "")

        if len(memberList) == 0:
            return render(request, 'motherwise/result.html',
                          {'response': 'The members don\'t exist.'})

        if len(memberList) > 0:
            groups = Group.objects.filter(member_id=admin.pk).order_by('-id')
            for group in groups:
                group.created_time = datetime.datetime.fromtimestamp(float(int(group.created_time)/1000)).strftime("%b %d, %Y %H:%M")
                group.last_connected_time = datetime.datetime.fromtimestamp(float(int(group.last_connected_time)/1000)).strftime("%b %d, %Y %H:%M")
            latestGroupMemberList = []
            latest_group = None
            if groups.count() > 0:
                latest_group = groups[0]
                gMembers = GroupMember.objects.filter(group_id=latest_group.pk)
                for gMember in gMembers:
                    members = Member.objects.filter(id=gMember.member_id)
                    if members.count() > 0:
                        latestGroupMemberList.append(members[0])
            gcs = GroupConnect.objects.filter(member_id=admin.pk).order_by('-id')
            recents = []
            for gc in gcs:
                gs = Group.objects.filter(id=gc.group_id)
                if gs.count() > 0:
                    group = gs[0]
                    group.created_time = datetime.datetime.fromtimestamp(float(int(group.created_time)/1000)).strftime("%b %d, %Y %H:%M")
                    group.last_connected_time = datetime.datetime.fromtimestamp(float(int(group.last_connected_time)/1000)).strftime("%b %d, %Y %H:%M")
                    recents.append(group)
            return render(request, 'motherwise/groups.html', {'members':memberList, 'group':latest_group, 'groups': groups, 'group_members':latestGroupMemberList, 'recents':recents})
        else:
            return redirect('/home')


def new_notis(request):
    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    unreadNotiList = []
    notis = Received.objects.filter(member_id=admin.pk)
    for noti in notis:
        nfs = Notification.objects.filter(id=noti.noti_id)
        if nfs.count() > 0:
            notification = nfs[0]
            if notification.status == '':
                unreadNotiList.append(notification)

    return HttpResponse(len(unreadNotiList))



@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def send_reply_message(request):
    if request.method == 'POST':
        member_id = request.POST.get('member_id', '1')
        noti_id = request.POST.get('noti_id', '1')
        message = request.POST.get('message', '')

        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        members = Member.objects.filter(id=int(member_id))
        if members.count() > 0:
            member = members[0]

            # title = 'You\'ve received a message from VaCay Community'
            # subject = 'From VaCay Community'
            # msg = 'Dear ' + member.name + ',<br><br>'
            # msg = msg + message

            # from_email = settings.ADMIN_EMAIL
            # to_emails = []
            # to_emails.append(member.email)
            # send_mail_message(from_email, to_emails, title, subject, msg)

            msg = member.name + ', You\'ve received a reply message from VaCay Community.\nThe message is as following:\n' + message

            notification = Notification()
            notification.member_id = member.pk
            notification.sender_id = admin.pk
            notification.message = msg
            notification.notified_time = str(int(round(time.time() * 1000)))
            notification.save()

            rcv = Received()
            rcv.member_id = member.pk
            rcv.sender_id = admin.pk
            rcv.noti_id = notification.pk
            rcv.save()

            snt = Sent()
            snt.member_id = member.pk
            snt.sender_id = admin.pk
            snt.noti_id = notification.pk
            snt.save()

            replieds = Replied.objects.filter(noti_id=noti_id)
            if replieds.count() == 0:
                nfs = Notification.objects.filter(id=noti_id)
                if nfs.count() > 0:
                    noti = nfs[0]
                    replied = Replied()
                    replied.root_id = noti.pk
                    replied.noti_id = noti.pk
                    replied.save()

                    replied = Replied()
                    replied.root_id = noti.pk
                    replied.noti_id = notification.pk
                    replied.save()
            else:
                repl = replieds[0]
                replied = Replied()
                replied.root_id = repl.root_id
                replied.noti_id = notification.pk
                replied.save()

            ##########################################################################################################################################################################

            db = firebase.database()
            data = {
                "msg": msg,
                "date":str(int(round(time.time() * 1000))),
                "sender_id": str(admin.pk),
                "sender_name": admin.name,
                "sender_email": admin.email,
                "sender_photo": admin.photo_url,
                "role": "admin",
                "type": "message",
                "id": str(notification.pk),
                "mes_id": str(notification.pk)
            }

            db.child("notify").child(str(member.pk)).push(data)
            db.child("notify2").child(str(member.pk)).push(data)

            sendFCMPushNotification(member.pk, admin.pk, msg)

            #################################################################################################################################################################################

            if member.playerID != '':
                playerIDList = []
                playerIDList.append(member.playerID)
                msg = member.name + ', You\'ve received a reply message from VaCay Community.\nThe message is as following:\n' + message
                url = '/users/notifications?noti_id=' + str(notification.pk)
                send_push(playerIDList, msg, url)

            return redirect('/notifications/')
        else:
            return render(request, 'motherwise/result.html',
                          {'response': 'This member doesn\'t exist.'})


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def send_member_message(request):
    if request.method == 'POST':
        member_id = request.POST.get('member_id', '1')
        message = request.POST.get('message', '')

        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        members = Member.objects.filter(id=int(member_id))
        if members.count() > 0:
            member = members[0]

            title = 'You\'ve received a message from VaCay Community'
            subject = 'VaCay Community'
            msg = 'Dear ' + member.name + ',<br><br>'
            msg = msg + message

            from_email = admin.email
            to_emails = []
            to_emails.append(member.email)
            send_mail_message(from_email, to_emails, title, subject, msg)

            msg = member.name + ', You\'ve received a message from VaCay Community.\nThe message is as following:\n' + message

            notification = Notification()
            notification.member_id = member.pk
            notification.sender_id = admin.pk
            notification.message = msg
            notification.notified_time = str(int(round(time.time() * 1000)))
            notification.save()

            rcv = Received()
            rcv.member_id = member.pk
            rcv.sender_id = admin.pk
            rcv.noti_id = notification.pk
            rcv.save()

            snt = Sent()
            snt.member_id = member.pk
            snt.sender_id = admin.pk
            snt.noti_id = notification.pk
            snt.save()

            ##########################################################################################################################################################################

            db = firebase.database()
            data = {
                "msg": message,
                "date":str(int(round(time.time() * 1000))),
                "sender_id": str(admin.pk),
                "sender_name": admin.name,
                "sender_email": admin.email,
                "sender_photo": admin.photo_url,
                "role": "admin",
                "type": "message",
                "id": str(notification.pk),
                "mes_id": str(notification.pk)
            }

            db.child("notify").child(str(member.pk)).push(data)
            db.child("notify2").child(str(member.pk)).push(data)

            sendFCMPushNotification(member.pk, admin.pk, message)

            #################################################################################################################################################################################

            if member.playerID != '':
                playerIDList = []
                playerIDList.append(member.playerID)
                msg = member.name + ', You\'ve received a message from VaCay Community.\nThe message is as following:\n' + message
                url = '/users/notifications?noti_id=' + str(notification.pk)
                send_push(playerIDList, msg, url)

            return redirect('/notifications/')
        else:
            return render(request, 'motherwise/result.html',
                          {'response': 'This member doesn\'t exist.'})



def to_conferences(request):

    import datetime

    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    confs = Conference.objects.filter(member_id=admin.pk).order_by('-id')
    gconfList = []
    cconfList = []
    for conf in confs:
        conf.created_time = datetime.datetime.fromtimestamp(float(int(conf.created_time)/1000)).strftime("%b %d, %Y %H:%M")
        if conf.event_time != '': conf.event_time = datetime.datetime.fromtimestamp(float(int(conf.event_time)/1000)).strftime("%b %d, %Y %H:%M")
        if conf.cohort == '':
            groups = Group.objects.filter(id=conf.group_id)
            if groups.count() > 0:
                group = groups[0]
                data={
                    'conf':conf,
                    'group': group
                }
                gconfList.append(data)
        elif conf.group_id == '':
            cconfList.append(conf)

    groups = Group.objects.filter(member_id=admin.pk).order_by('-id')

    return render(request, 'motherwise/conferences.html', {'conf_opt':'new_conference', 'gconfs':gconfList, 'cconfs':cconfList, 'groups':groups})




from twilio.rest import Client

def sendSMS(to_phone, msg):

    # Your Account Sid and Auth Token from twilio.com/console
    # DANGER! This is insecure. See http://twil.io/secure
    account_sid = 'ACa84d7b1bddaec4ba6465060ae44fb2f3'
    auth_token = 'bfc5cdda6bf320a153116fd80b2a9b7a'
    client = Client(account_sid, auth_token)

    message = client.messages \
                    .create(
                         body=msg,
                         from_='+17206795056',
                         to=to_phone
                     )

    print(message.sid)

    return to_phone


def sms_test(request):
    to_phone = sendSMS('+18438161828', '12345')         #  18438161828
    return HttpResponse('SMS sent to ' + to_phone)


def noti_detail(request):
    import datetime

    noti_id = request.GET['noti_id']
    opt = request.GET['opt']

    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    list = []
    sender0 = None

    replieds = Replied.objects.filter(noti_id=noti_id)
    if replieds.count() > 0:
        repl = replieds[0]
        repls = Replied.objects.filter(root_id=repl.root_id)
        for repl in repls:
            notis = Notification.objects.filter(id=repl.noti_id)
            if notis.count() > 0:
                noti = notis[0]
                date = datetime.datetime.fromtimestamp(float(int(noti.notified_time)/1000)).strftime("%b %d, %Y %H:%M")
                members = Member.objects.filter(id=noti.sender_id)
                if members.count() > 0:
                    sender = members[0]
                    if sender.pk != admin.pk:
                        sender0 = sender
                    data = {
                        'sender':sender,
                        'noti': noti,
                        'date':date
                    }

                    list.append(data)

    else:
        notis = Notification.objects.filter(id=noti_id)
        if notis.count() > 0:
            noti = notis[0]
            date = datetime.datetime.fromtimestamp(float(int(noti.notified_time)/1000)).strftime("%b %d, %Y %H:%M")
            members = Member.objects.filter(id=noti.sender_id)
            if members.count() > 0:
                sender = members[0]
                sender0 = sender
                data = {
                    'sender':sender,
                    'noti': noti,
                    'date':date
                }

                list.append(data)

    return render(request, 'motherwise/noti_detail.html', {'notid':noti_id, 'me':admin, 'sender':sender0, 'list':list, 'opt':opt})



@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def notify_group_chat(request):
    if request.method == 'POST':

        message = request.POST.get('message', '')
        cohort = request.POST.get('cohort', '')
        groupid = request.POST.get('groupid', '')
        ids = request.POST.getlist('users[]')

        try:
            if request.session['adminID'] == '' or request.session['adminID'] == 0:
                return render(request, 'motherwise/admin.html')
        except KeyError:
            print('no session')
            return render(request, 'motherwise/admin.html')

        adminID = request.session['adminID']
        admin = Member.objects.get(id=adminID)

        if groupid != '':
            groups = Group.objects.filter(id=int(groupid))
            if groups.count() == 0:
                return redirect('/home')

            group = groups[0]

            for member_id in ids:
                members = Member.objects.filter(id=int(member_id))
                if members.count() > 0:
                    member = members[0]

                    title = 'VaCay Community'
                    subject = 'You\'ve received a community message from ' + group.name
                    msg = 'Dear ' + member.name + ', You\'ve received a community message from manager in ' + group.name + '. The message is as following:<br><br>'
                    msg = msg + message + '<br><br>'
                    msg = msg + '<a href=\'' + settings.URL + '/users/open_group_chat?group_id=' + groupid + '\' target=\'_blank\'>Connect the community to view message</a>'

                    from_email = admin.email
                    to_emails = []
                    to_emails.append(member.email)
                    send_mail_message(from_email, to_emails, title, subject, msg)

                    msg = member.name + ', You\'ve received a community message from manager in ' + group.name + '. The message is as following:\n\n'
                    msg = msg + message + '\n\n'
                    msg = msg + 'Click on this link to view the message: ' + settings.URL + '/users/open_group_chat?group_id=' + groupid

                    notification = Notification()
                    notification.member_id = member.pk
                    notification.sender_id = admin.pk
                    notification.message = msg
                    notification.notified_time = str(int(round(time.time() * 1000)))
                    notification.save()

                    rcv = Received()
                    rcv.member_id = member.pk
                    rcv.sender_id = admin.pk
                    rcv.noti_id = notification.pk
                    rcv.save()

                    snt = Sent()
                    snt.member_id = member.pk
                    snt.sender_id = admin.pk
                    snt.noti_id = notification.pk
                    snt.save()

                    ##########################################################################################################################################################################

                    db = firebase.database()
                    data = {
                        "msg": msg,
                        "date":str(int(round(time.time() * 1000))),
                        "sender_id": str(admin.pk),
                        "sender_name": admin.name,
                        "sender_email": admin.email,
                        "sender_photo": admin.photo_url,
                        "role": "admin",
                        "type": "group_chat",
                        "id": str(group.pk),
                        "mes_id": str(notification.pk)
                    }

                    db.child("notify").child(str(member.pk)).push(data)
                    db.child("notify2").child(str(member.pk)).push(data)

                    sendFCMPushNotification(member.pk, admin.pk, msg)

                    #################################################################################################################################################################################

                    if member.playerID != '':
                        playerIDList = []
                        playerIDList.append(member.playerID)
                        url = '/users/notifications?noti_id=' + str(notification.pk)
                        send_push(playerIDList, msg, url)

        elif cohort != '':

            for member_id in ids:
                members = Member.objects.filter(id=int(member_id))
                if members.count() > 0:
                    member = members[0]

                    title = 'VaCay Community'
                    subject = 'You\'ve received a group message from ' + cohort
                    msg = 'Dear ' + member.name + ', You\'ve received a group message from manager in ' + cohort + '. The message is as following:<br><br>'
                    msg = msg + message + '<br><br>'
                    msg = msg + '<a href=\'' + settings.URL + '/users/open_cohort_chat?cohort=' + cohort + '\' target=\'_blank\'>Connect the group to view message</a>'

                    from_email = admin.email
                    to_emails = []
                    to_emails.append(member.email)
                    send_mail_message(from_email, to_emails, title, subject, msg)

                    msg = member.name + ', You\'ve received a group message from manager in ' + cohort + '. The message is as following:\n\n'
                    msg = msg + message + '\n\n'
                    msg = msg + 'Click on this link to view the message: ' + settings.URL + '/users/open_cohort_chat?cohort=' + cohort

                    notification = Notification()
                    notification.member_id = member.pk
                    notification.sender_id = admin.pk
                    notification.message = msg
                    notification.notified_time = str(int(round(time.time() * 1000)))
                    notification.save()

                    rcv = Received()
                    rcv.member_id = member.pk
                    rcv.sender_id = admin.pk
                    rcv.noti_id = notification.pk
                    rcv.save()

                    snt = Sent()
                    snt.member_id = member.pk
                    snt.sender_id = admin.pk
                    snt.noti_id = notification.pk
                    snt.save()

                    ##########################################################################################################################################################################

                    db = firebase.database()
                    data = {
                        "msg": msg,
                        "date":str(int(round(time.time() * 1000))),
                        "sender_id": str(admin.pk),
                        "sender_name": admin.name,
                        "sender_email": admin.email,
                        "sender_photo": admin.photo_url,
                        "role": "admin",
                        "type": "cohort_chat",
                        "id": cohort,
                        "mes_id": str(notification.pk)
                    }

                    db.child("notify").child(str(member.pk)).push(data)
                    db.child("notify2").child(str(member.pk)).push(data)

                    sendFCMPushNotification(member.pk, admin.pk, msg)

                    #################################################################################################################################################################################

                    if member.playerID != '':
                        playerIDList = []
                        playerIDList.append(member.playerID)
                        url = '/users/notifications?noti_id=' + str(notification.pk)
                        send_push(playerIDList, msg, url)

        return HttpResponse('success')


def open_translate(request):
    return render(request, 'motherwise/translate.html')

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def process_translate(request):
    from googletrans import Translator
    if request.method == 'POST':
        input = request.POST.get('input', '')
        lang = request.POST.get('lang', '')

        translator = Translator()
        translation = translator.translate(input, dest=lang)
        return HttpResponse(translation.text)



def analytics(request):

    from datetime import datetime

    try:
        if request.session['adminID'] == '' or request.session['adminID'] == 0:
            return render(request, 'motherwise/admin.html')
    except KeyError:
        print('no session')
        return render(request, 'motherwise/admin.html')

    adminID = request.session['adminID']
    admin = Member.objects.get(id=adminID)

    members = Member.objects.filter(admin_id=admin.pk)
    groups = ['E81', 'E83', 'E84', 'E86', 'E87', 'S82', 'S85', 'S88', 'E(v)89', 'E(v)90', 'S(v)91', 'E(v)92', 'E(v)93', 'S(v)94', 'VACAY Leaders']

    inviteds = [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
    activateds = [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
    group_posts = [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
    total_activs = 0
    activ_percentlist_bygroup = []
    monthly_activateds = [0,0,0,0,0,0,0,0,0,0,0,0]
    monthly_posts = [0,0,0,0,0,0,0,0,0,0,0,0]

    cities = []
    activateds_bycity = []
    activated_members_bycity = []

    nowObj = datetime.fromtimestamp(time.time())

    for member in members:
        if member.cohort != 'admin' and member.cohort != '':
            index_invited = groups.index(member.cohort)
            invits = inviteds[index_invited] + 1
            inviteds[index_invited] = invits
            if len(member.registered_time) > 0 and int(member.registered_time) > 0:
                index_activated = groups.index(member.cohort)
                activs = activateds[index_activated] + 1
                activateds[index_activated] = activs
                total_activs = total_activs + 1

                posts = Post.objects.filter(member_id=member.pk)
                for post in posts:
                    psts = group_posts[index_activated] + 1
                    group_posts[index_activated] = psts

        if len(member.registered_time) > 0 and int(member.registered_time) > 0:
            regDateObj = datetime.fromtimestamp(int(member.registered_time)/1000)
            if nowObj.year == regDateObj.year:
                activs = monthly_activateds[regDateObj.month - 1] + 1
                monthly_activateds[regDateObj.month - 1] = activs
            if member.city not in cities:
                cities.append(member.city.replace('\'',''))
                activateds_bycity.append(0)
                activated_members_bycity.append([])
            activs = activateds_bycity[cities.index(member.city)] + 1
            activateds_bycity[cities.index(member.city)] = activs
            activated_members_bycity[cities.index(member.city)].append(member)

            posts = Post.objects.filter(member_id=member.pk)
            for post in posts:
                postDateObj = datetime.fromtimestamp(int(post.posted_time)/1000)
                if nowObj.year == postDateObj.year:
                    psts = monthly_posts[postDateObj.month - 1] + 1
                    monthly_posts[postDateObj.month - 1] = psts


    cityActivsList = []
    cityActivsChartWidth = 0
    for i in range(0, len(cities)):
        data = {
            'city': cities[i],
            'activsval': activateds_bycity[i],
            'activmembers': activated_members_bycity[i]
        }
        cityActivsList.append(data)
        cityActivsChartWidth = cityActivsChartWidth + 50

    for i in range(0, len(activateds)):
        percentval = round(activateds[i] * 100 / total_activs, 2)
        data = {
            'group': groups[i],
            'activ_percent': percentval
        }
        activ_percentlist_bygroup.append(data)

    activ_percent = round(total_activs * 100 / members.count(), 2)
    activdata = {
        'percvalue': activ_percent,
        'title': 'Activated'
    }
    inactivdata = {
        'percvalue': round(100 - activ_percent, 2),
        'title': 'Inactivated'
    }

    context = {
        'inviteds': inviteds,
        'activateds': activateds,
        'group_posts': group_posts,
        'activ_percentlist_bygroup': activ_percentlist_bygroup,
        'total_activ_inactiv_percent': [activdata, inactivdata],
        'monthly_activateds': monthly_activateds,
        'monthly_posts': monthly_posts,
        'city_activateds': cityActivsList,
        'cityActivsChartWidth': cityActivsChartWidth
    }

    return render(request, 'motherwise/analytics.html', context)



@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def sendfcmpush(request):
    if request.method == 'POST':

        sender_id = request.POST.get('sender_id', '1')
        receiver_id = request.POST.get('receiver_id', '1')
        message = request.POST.get('message', '')

        sendFCMPushNotification(receiver_id, sender_id, message)

        senders = Member.objects.filter(id=sender_id)
        if senders.count() == 0:
            resp = {'result_code': '1'}
            return HttpResponse(json.dumps(resp))

        sender = senders[0]

        members = Member.objects.filter(id=receiver_id)
        if members.count() > 0:
            member = members[0]
            if member.playerID != '':
                playerIDList = []
                playerIDList.append(member.playerID)
                url = '/users/to_private_chat?member_id=' + str(sender.pk)
                if member.cohort == 'admin':
                    url = '/group_private_chat?email=' + sender.email
                msg = member.name + ', You\'ve received a message from ' + sender.name + '.\nThe message is as following:\n' + message
                send_push(playerIDList, msg, url)

        resp = {'result_code': '0'}
        return HttpResponse(json.dumps(resp))



def sendFCMPushNotification(member_id, sender_id, notiText):
    members = Member.objects.filter(id=member_id)
    if members.count() > 0:
        member = members[0]
        message_title = 'VaCay User'
        if int(sender_id) > 0:
            senders = Member.objects.filter(id=sender_id)
            if senders.count() > 0:
                sender = senders[0]
                message_title = sender.name
        path_to_fcm = "https://fcm.googleapis.com"
        server_key = settings.FCM_LEGACY_SERVER_KEY
        reg_id = member.fcm_token #quick and dirty way to get that ONE fcmId from table
        if reg_id != '':
            message_body = notiText
            result = FCMNotification(api_key=server_key).notify_single_device(registration_id=reg_id, message_title=message_title, message_body=message_body, sound = 'ping.aiff', badge = 1)



def landing(request):
    return render(request, 'landing/splash.html')

def main(request):
    return render(request, 'landing/main.html')


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def landing_notify(request):
    if request.method == 'POST':

        first_name = request.POST.get('first_name', '')
        last_name = request.POST.get('last_name', '')
        email = request.POST.get('email', '')

        members = Member.objects.filter(email=email)
        if members.count() > 0:
            return render(request, 'motherwise/result.html',
                          {'response': 'You\'ve already registered in VaCay platform.'})

        subject = 'VaCay-interested person\'s request of invitation'
        title = 'I\'m interested in VaCay'
        message = 'Hi, my name is ' + first_name + ' ' + last_name + ' and email is ' + email + '.\nI\'m interested in VaCay platform and I\'ll be happy if I receive your invitation to VaCay.\nRegards'

        from_email = email
        to_emails = []
        to_emails.append('cayley@vacaycarpediem.com')
        send_mail_message(from_email, to_emails, title, subject, '')

        return render(request, 'motherwise/result.html',
                          {'response': 'Your submission done successfully. Please wait for our reply.'})
















































def change(request):

    ### Member
    members = Member.objects.all()
    for member in members:
        member.email = encrypt(member.email)
        member.password = encrypt(member.password)
        if member.photo_url != '':
            member.photo_url = encrypt(member.photo_url)
        if member.phone_number != '':
            member.phone_number = encrypt(member.phone_number)

        if member.city != '':
            member.city = encrypt(member.city)
        if member.address != '':
            member.address = encrypt(member.address)

        member.save()

    ### Conference
    confs = Conference.objects.all()
    for conf in confs:
        conf.name = encrypt(conf.name)
        if conf.video_url != '':
            conf.video_url = encrypt(conf.video_url)
        conf.save()

    ### Contact
    contacts = Contact.objects.all()
    for contact in contacts:
        contact.contact_email = encrypt(contact.contact_email)
        contact.save()

    ### Group
    groups = Group.objects.all()
    for group in groups:
        group.name = encrypt(group.name)
        group.save()

    ### Notification
    notis = Notification.objects.all()
    for noti in notis:
        noti.message = encrypt(noti.message)
        noti.save()

    ### Post
    posts = Post.objects.all()
    for post in posts:
        post.title = encrypt(post.title)
        post.category = encrypt(post.category)
        post.content = encrypt(post.content)
        if post.picture_url != '':
            post.picture_url = encrypt(post.picture_url)
        if post.video_url != '':
            post.video_url = encrypt(post.video_url)
        post.save()

    ### Post Picture
    pps = PostPicture.objects.all()
    for pp in pps:
        pp.picture_url = encrypt(pp.picture_url)
        pp.save()



































